"""Smoke test - Prompt 1 structural rework."""
import sys, os, math
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from assembly.engine import run, ProjectConfig, TimelineConfig, TaxConfig, StatementConfig
from assembly.excel_writer import write_workbook

from modules.core.WACC_001 import Inputs as WACCInputs
from modules.revenue.REV_001 import Inputs as REVInputs
from modules.costs.OPEX_001 import Inputs as OPEXInputs, ComponentRate
from modules.capex.CAPEX_001 import Inputs as CAPEXInputs
from modules.debt.DEBT_001 import Inputs as DEBTInputs

# Full 20-year model
n = 240
cp = 8  # construction months

config = ProjectConfig(
    project_name="Format Test - DK Solar PV Full",
    market="DK",
    timeline=TimelineConfig(periods=n, start_year=2026, start_month=1),
    wacc=WACCInputs(
        country="DK",
        ppa_share=0.70,
        debt_aud=250_000.0,
        equity_aud=100_000.0,
    ),
    rev_pv=REVInputs(
        periods=n, start_year=2026, start_month=1,
        net_production_MWh=[0.0] * cp + [3500.0] * (n - cp),
        wholesale_price_DKK_per_MWh=[400.0] * n,
        capture_price_DKK_per_MWh=[380.0] * n,
        goo_price_DKK_per_MWh=[5.0] * n,
    ),
    opex_pv=OPEXInputs(
        periods=n, start_year=2026, start_month=1,
        capacity_mwp=50.0,
        om=ComponentRate(annual_DKKk=2500.0),
        commercial_management=ComponentRate(annual_DKKk=0.0),
        inverter_replacement=ComponentRate(annual_DKKk=0.0),
        insurance=ComponentRate(annual_DKKk=500.0),
        other=ComponentRate(annual_DKKk=0.0),
    ),
    capex=CAPEXInputs(
        periods=n,
        construction_start_period=0, construction_periods=cp,
        drawdown_profile=[1.0 / cp] * cp,
        epc_DKKk=280_000.0, grid_connection_DKKk=30_000.0,
        development_DKKk=20_000.0, land_DKKk=0.0,
        contingency_pct=0.05, other_DKKk=0.0,
    ),
    debt=DEBTInputs(
        periods=n, start_year=2026, start_month=1,
        facility=250_000.0, all_in_rate=0.055,
        repayment_type="annuity", tenor_months=180,
        drawdowns=[250_000.0 / cp] * cp + [0.0] * (n - cp),
    ),
    tax=TaxConfig(country="DK"),
    statements=StatementConfig(
        opening_cash_DKKk=1_000.0,
        opening_contributed_equity_DKKk=100_000.0,
    ),
)

os.makedirs("outputs", exist_ok=True)
result = run(config)
write_workbook(result, config, "outputs/format_smoke_test.xlsx")
print(f"Written: outputs/format_smoke_test.xlsx  |  Warnings: {len(result.warnings)}")

import openpyxl
wb = openpyxl.load_workbook("outputs/format_smoke_test.xlsx")

assert "FS_Monthly"  in wb.sheetnames, f"FAIL: FS_Monthly missing. Got: {wb.sheetnames}"
assert "FS_Annual"   in wb.sheetnames, f"FAIL: FS_Annual missing. Got: {wb.sheetnames}"
assert "Cover"       in wb.sheetnames, f"FAIL: Cover missing. Got: {wb.sheetnames}"
assert "Statements" not in wb.sheetnames, "FAIL: old Statements sheet still present"

ws = wb["Revenue"]
assert ws.cell(row=5, column=5).value == "Description", \
    f"FAIL: E5 = '{ws.cell(row=5,column=5).value}', expected 'Description'"
# Check that data exists somewhere in the Revenue sheet below the header
has_data = any(
    ws.cell(row=r, column=12).value is not None
    for r in range(7, min(ws.max_row + 1, 30))
)
assert has_data, "FAIL: no data found in col L rows 7-30 on Revenue sheet"
assert str(ws.freeze_panes) == "L7", \
    f"FAIL: freeze pane = '{ws.freeze_panes}', expected 'L7'"
assert ws.column_dimensions.get("H") and ws.column_dimensions["H"].hidden, \
    "FAIL: col H should be hidden"
assert ws.column_dimensions.get("I") and ws.column_dimensions["I"].hidden, \
    "FAIL: col I should be hidden"

fsa = wb["FS_Annual"]
yr_first = fsa.cell(row=2, column=12).value
assert isinstance(yr_first, int) and yr_first == 2026, \
    f"FAIL: FS_Annual col L should be 2026, got {yr_first!r}"

# 240 months from Jan 2026 = 20 full years (2026-2045) -> col 31 = 2045
yr_last = fsa.cell(row=2, column=31).value
assert isinstance(yr_last, int) and yr_last == 2045, \
    f"FAIL: FS_Annual col 31 should be 2045, got {yr_last!r}"

print("All structural assertions PASSED.")
