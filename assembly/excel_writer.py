"""
assembly/excel_writer.py

Writes an AssemblyResult to a bank-grade .xlsx workbook using openpyxl.

Sheet layout (EE Standard):
    "Cover"       — project name stub
    "Revenue"     — REV_001, REV_002, REV_003, PPA_CFD_001
    "Costs"       — OPEX_001, OPEX_002, OPEX_003, CAPEX_001, BESS_REPOW_001
    "Debt"        — DEBT_001, TAX_001, WACC_001 (scalars in col F)
    "FS_Monthly"  — PL_001, CF_001, BS_001, IRR_001 (monthly time series)
    "FS_Annual"   — same modules, annual aggregation
    "Summary"     — scalar KPIs

Column layout (EE Standard):
    A-D  = narrow indent spacers (width 1.3)
    E    = row description / label (width 40.5)
    F    = constant / assumption value (width 12.5)
    G    = unit (width 14.5)
    H    = notes (hidden, outline 1) (width 45.5)
    I    = source (hidden, outline 1) (width 45.5)
    J    = total / lifetime avg (width 15.5)
    K    = spacer (width 2.5)
    L+   = time series (width 11.5 each), period 0 = col L
"""

from __future__ import annotations

import math
import re
from datetime import date, timedelta
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from assembly.cell_mapper import (
    ROW_MAP,
    MODULE_SHEET,
    COL_LABEL,
    COL_CONSTANT,
    COL_UNIT,
    COL_TOTAL,
    COL_PERIOD_0,
    COL_WIDTHS,
    HIDDEN_COLS,
    get_label,
    get_units,
    period_col,
    col_letter,
)
from assembly.engine import AssemblyResult, ProjectConfig
from assembly.excel_formatter import apply_formatting
from assembly.cover_writer import write_cover

import modules.statements.PL_001 as PL_001
import modules.statements.CF_001 as CF_001
import modules.statements.BS_001 as BS_001

# ============================================================================
# CONSTANTS
# ============================================================================

SHEETS = ["Cover", "Revenue", "Costs", "Debt", "FS_Monthly", "FS_Annual", "Summary"]

# WACC_001 scalar fields — written to col F on Debt sheet (no time-series columns)
WACC_SCALAR_FIELDS = [
    "wacc",
    "blended_cost_of_equity",
    "ppa_cost_of_equity",
    "merchant_cost_of_equity",
    "cost_of_debt_posttax",
    "levered_beta",
    "debt_to_equity_ratio",
    "ppa_erp",
    "merchant_erp",
]

# IRR_001 scalar fields → Summary tab
IRR_SCALAR_FIELDS = [
    "project_irr",
    "equity_irr",
    "project_npv",
    "equity_npv",
    "payback_period_months",
]

# Number formats
FMT_DKKK    = '#,##0.00'
FMT_PCT     = '0.00%'
FMT_RATIO   = '0.000'
FMT_INTEGER = '#,##0'
FMT_GENERAL = 'General'

_PCT_FIELDS   = {"wacc", "blended_cost_of_equity", "ppa_cost_of_equity",
                 "merchant_cost_of_equity", "cost_of_debt_posttax",
                 "ppa_erp", "merchant_erp", "project_irr", "equity_irr"}
_RATIO_FIELDS = {"levered_beta", "debt_to_equity_ratio", "dscr_annual",
                 "price_indexation_factor", "discharge_indexation_factor",
                 "charge_indexation_factor", "dscr_achieved", "dscr_covenant_applicable"}
_INT_FIELDS   = {"payback_period_months"}
_MWH_FIELDS   = {"net_production_MWh", "discharge_volume", "goo_volume",
                 "goo_lost_volume", "total_ppa_contracted_volume"}

_BOLD_FONT    = Font(bold=True)


# ============================================================================
# HELPERS
# ============================================================================

_MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _year_groups(n: int, start_year: int, start_month: int) -> dict:
    """Group period indices (0-indexed) by calendar year.
    Returns {year: [period_indices]}.
    """
    groups = {}
    for p in range(n):
        year = start_year + ((start_month - 1) + p) // 12
        groups.setdefault(year, []).append(p)
    return groups


def _period_to_label(start_year: int, start_month: int, period: int) -> str:
    """Return 'Jan-2026' label for period p given project start."""
    total_months = (start_month - 1) + period
    year  = start_year + total_months // 12
    month = total_months % 12
    return f"{_MONTH_NAMES[month]}-{year}"


def _period_end_date(start_year: int, start_month: int, period: int) -> date:
    """Return the last day of the month for the given period."""
    total_months = (start_month - 1) + period
    year = start_year + total_months // 12
    month = total_months % 12 + 1  # 1-based month
    # Last day: go to 1st of next month, subtract 1 day
    if month == 12:
        return date(year, 12, 31)
    else:
        return date(year, month + 1, 1) - timedelta(days=1)


def _safe(value: Any) -> Any:
    """Replace NaN/inf with empty string so Excel accepts the cell value."""
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return ""
    return value


def _num_format(field_name: str) -> str:
    if field_name in _PCT_FIELDS:
        return FMT_PCT
    if field_name in _RATIO_FIELDS:
        return FMT_RATIO
    if field_name in _INT_FIELDS:
        return FMT_INTEGER
    return FMT_DKKK


def _get_field_value(module_out: Any, field_name: str) -> Any:
    """Return the attribute named field_name from a module output object."""
    return getattr(module_out, field_name, None)


# ============================================================================
# HEADER ROWS
# ============================================================================

def _write_header_rows(
    ws, sheet_name: str, n: int,
    start_year: int, start_month: int,
    capex_monthly=None, annual_mode: bool = False,
) -> None:
    """Write rows 1-6: title, period end dates, phase labels, year, column headers, blank."""
    yg = _year_groups(n if not annual_mode else n * 12, start_year, start_month) \
        if not annual_mode else None

    # Row 1: sheet name in col E, bold
    ws.cell(row=1, column=COL_LABEL, value=sheet_name).font = _BOLD_FONT

    if annual_mode:
        # Annual mode: n = number of years, derive year list from full periods
        # We compute year_groups from the caller's context
        # Here n = n_years, and we need the actual year integers
        # The caller passes n_years as n; we reconstruct years from start
        years = []
        y = start_year
        sm = start_month
        # Build year list by walking months
        month_counter = sm - 1
        seen = {}
        # Actually simpler: use _year_groups on a synthetic full period count
        # But we don't know total months here. Instead, use the fact that
        # col L+ has n columns, one per year. Row 2 = year integers.
        for i in range(n):
            # Approximate: first year may be partial
            yr = start_year + i if start_month == 1 else start_year + i
            # More precise: derive from _year_groups if we had the month count
            years.append(start_year + i)

        # Row 2: year integers
        for i, yr in enumerate(years):
            ws.cell(row=2, column=COL_PERIOD_0 + i, value=yr)

        # Row 3: phase labels — phase of LAST month in each year group
        # We need capex_monthly which has monthly granularity
        if capex_monthly is not None:
            total_months = len(capex_monthly)
            monthly_yg = _year_groups(total_months, start_year, start_month)
            for i, (year, periods) in enumerate(monthly_yg.items()):
                last_p = periods[-1]
                if last_p < len(capex_monthly) and capex_monthly[last_p] > 0:
                    ws.cell(row=3, column=COL_PERIOD_0 + i, value="Construction")
                else:
                    ws.cell(row=3, column=COL_PERIOD_0 + i, value="Operations")
        else:
            for i in range(n):
                ws.cell(row=3, column=COL_PERIOD_0 + i, value="Operations")

        # Row 4: calendar year integer (same as row 2 for annual)
        for i, yr in enumerate(years):
            ws.cell(row=4, column=COL_PERIOD_0 + i, value=yr)

        # Row 5: column headers
        ws.cell(row=5, column=COL_LABEL, value="Description").font = _BOLD_FONT
        ws.cell(row=5, column=COL_CONSTANT, value="Constant").font = _BOLD_FONT
        ws.cell(row=5, column=COL_UNIT, value="Unit").font = _BOLD_FONT
        ws.cell(row=5, column=COL_TOTAL, value="Total / avg.").font = _BOLD_FONT
        for i, yr in enumerate(years):
            ws.cell(row=5, column=COL_PERIOD_0 + i, value=yr).font = _BOLD_FONT

    else:
        # Monthly mode
        # Row 2: period end dates as datetime objects
        for p in range(n):
            dt = _period_end_date(start_year, start_month, p)
            cell = ws.cell(row=2, column=COL_PERIOD_0 + p, value=dt)
            cell.number_format = "DD-MMM-YY"

        # Row 3: phase labels
        for p in range(n):
            if capex_monthly is not None and p < len(capex_monthly) and capex_monthly[p] > 0:
                ws.cell(row=3, column=COL_PERIOD_0 + p, value="Construction")
            else:
                ws.cell(row=3, column=COL_PERIOD_0 + p, value="Operations")

        # Row 4: calendar year integer
        for p in range(n):
            year = start_year + ((start_month - 1) + p) // 12
            ws.cell(row=4, column=COL_PERIOD_0 + p, value=year)

        # Row 5: column headers
        ws.cell(row=5, column=COL_LABEL, value="Description").font = _BOLD_FONT
        ws.cell(row=5, column=COL_CONSTANT, value="Constant").font = _BOLD_FONT
        ws.cell(row=5, column=COL_UNIT, value="Unit").font = _BOLD_FONT
        ws.cell(row=5, column=COL_TOTAL, value="Total / avg.").font = _BOLD_FONT
        for p in range(n):
            ws.cell(row=5, column=COL_PERIOD_0 + p, value=p + 1).font = _BOLD_FONT

    # Row 6: blank spacer (leave empty)


# ============================================================================
# COLUMN FORMATTING
# ============================================================================

def _apply_column_formatting(ws, n_periods: int) -> None:
    """Set column widths and hide notes/source columns."""
    for col_idx, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for p in range(n_periods):
        ws.column_dimensions[get_column_letter(COL_PERIOD_0 + p)].width = 11.46
    for col_idx in HIDDEN_COLS:
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].hidden = True
        ws.column_dimensions[letter].outlineLevel = 1


# ============================================================================
# SUB-SECTION LABELS
# ============================================================================

def _write_subsection_labels(ws, sheet_name: str) -> None:
    """Write sub-section label text into pre-planned label rows.

    These rows sit between module sections and contain only a label
    in col E. They are formatted by excel_formatter.py separately.
    They must NOT conflict with ROW_MAP data rows.
    """
    from assembly.cell_mapper import SUBSECTION_LABELS

    for (sht, row_num), label in SUBSECTION_LABELS.items():
        if sht != sheet_name:
            continue
        ws.cell(row=row_num, column=COL_LABEL, value=label)


# ============================================================================
# SHEET WRITERS
# ============================================================================

def _write_time_series_rows(
    ws, result: AssemblyResult,
    assumption_map: dict[str, str] | None = None,
) -> None:
    """
    Write all ROW_MAP entries that belong to this worksheet as time-series rows.

    Two-layer strategy:
      Layer 1: If get_excel_formulas() provides a formula → write the formula.
      Layer 2: Otherwise → write the Python-computed value (fallback).

    Skips WACC_001 (handled separately) and rows for disabled modules.
    """
    sheet_name = ws.title
    # Map FS_Monthly -> Statements for ROW_MAP lookup
    lookup_sheet = "Statements" if sheet_name == "FS_Monthly" else sheet_name
    if assumption_map is None:
        assumption_map = {}

    entries = [
        (row, mid, field)
        for (sh, mid, field), row in ROW_MAP.items()
        if sh == lookup_sheet and mid != "WACC_001"
    ]
    entries.sort()

    for row_num, module_id, field_name in entries:
        module_out = result.outputs.get(module_id)

        # Label col E
        ws.cell(row=row_num, column=COL_LABEL, value=get_label(field_name)).font = _BOLD_FONT
        # Units col G
        ws.cell(row=row_num, column=COL_UNIT, value=get_units(field_name))

        if module_out is None:
            continue  # module disabled — leave data cells empty

        value = _get_field_value(module_out, field_name)
        if value is None:
            # Field not found on output — skip silently, add warning
            result.warnings.append(
                f"ROW_MAP field {module_id}.{field_name} not found on output"
            )
            continue

        fmt = _num_format(field_name)

        if isinstance(value, list):
            # Write time series to col L+
            for p, v in enumerate(value):
                # Layer 1: try formula
                formula_str = _try_get_formula(
                    module_id, field_name, p, result, assumption_map,
                )
                if formula_str is not None:
                    ws.cell(row=row_num, column=period_col(p),
                            value=formula_str).number_format = fmt
                else:
                    # Layer 2: value fallback
                    ws.cell(row=row_num, column=period_col(p),
                            value=_safe(v)).number_format = fmt

            # Write total/avg to col J
            non_zero = [v for v in value if isinstance(v, (int, float))
                        and not math.isnan(v) and v != 0]
            if field_name in _PCT_FIELDS | _RATIO_FIELDS:
                total_val = sum(non_zero) / len(non_zero) if non_zero else 0.0
            else:
                total_val = sum(_safe(v) or 0.0 for v in value)
            ws.cell(row=row_num, column=COL_TOTAL,
                    value=_safe(total_val)).number_format = fmt
        else:
            # Scalar on a time-series row — write to col F
            ws.cell(row=row_num, column=COL_CONSTANT,
                    value=_safe(value)).number_format = fmt


def _write_wacc_scalars(ws, result: AssemblyResult) -> None:
    """Write WACC_001 scalar outputs to col F of Debt sheet."""
    wacc_out = result.outputs.get("WACC_001")
    if wacc_out is None:
        return

    for field_name in WACC_SCALAR_FIELDS:
        row_num = ROW_MAP.get(("Debt", "WACC_001", field_name))
        if row_num is None:
            continue
        value = getattr(wacc_out, field_name, None)
        if value is None:
            continue
        ws.cell(row=row_num, column=COL_LABEL, value=get_label(field_name))
        ws.cell(row=row_num, column=COL_CONSTANT, value=_safe(value))
        ws.cell(row=row_num, column=COL_UNIT, value=get_units(field_name))
        ws.cell(row=row_num, column=COL_CONSTANT).number_format = _num_format(field_name)
        # No COL_TOTAL — WACC fields have no period list


def _write_fs_annual(ws, result: AssemblyResult, config: ProjectConfig) -> None:
    """Annual FS aggregations.

    Stocks (balance sheet): closing value of LAST period in year.
    Flows (P&L, CF):        SUM of all periods in year.
    """
    BS_CLOSING_FIELDS = {
        "fixed_assets_gross", "accumulated_depreciation", "fixed_assets_net",
        "cash", "total_assets", "debt_balance", "equity",
        "retained_earnings", "total_liabilities_equity",
        "contributed_equity", "total_equity",
    }
    n = result.periods
    sy = result.start_year
    sm = result.start_month
    year_groups = _year_groups(n, sy, sm)
    n_years = len(year_groups)

    capex_out = result.outputs.get("CAPEX_001")
    capex_monthly = capex_out.total_capex_monthly if capex_out else None
    _write_header_rows(ws, "FS_Annual", n_years, sy, sm,
                       capex_monthly=capex_monthly, annual_mode=True)

    # ROW_MAP uses sheet="Statements" for PL/CF/BS — reuse those row numbers
    for mod_id in ["PL_001", "CF_001", "BS_001", "IRR_001",
                   "WORKING_CAPITAL_001", "SOURCES_USES_001",
                   "VALUATION_001", "BREAKEVEN_001",
                   "MODEL_CHECKS_001", "DASHBOARD_001"]:
        mod_out = result.outputs.get(mod_id)
        if mod_out is None:
            continue
        for (sht, mid, field), row_num in ROW_MAP.items():
            if sht != "Statements" or mid != mod_id:
                continue
            values = getattr(mod_out, field, None)
            if not isinstance(values, list):
                continue

            ws.cell(row=row_num, column=COL_LABEL, value=get_label(field))
            ws.cell(row=row_num, column=COL_UNIT, value=get_units(field))

            annual_vals = []
            for year, periods in year_groups.items():
                pv = [values[p] for p in periods if p < len(values)]
                agg = pv[-1] if (field in BS_CLOSING_FIELDS and pv) \
                      else sum(_safe(v) or 0.0 for v in pv)
                annual_vals.append(agg)

            for i, v in enumerate(annual_vals):
                ws.cell(row=row_num, column=COL_PERIOD_0 + i,
                        value=_safe(v)).number_format = _num_format(field)

            non_zero = [v for v in annual_vals
                        if isinstance(v, (int, float)) and v != 0]
            if field in _PCT_FIELDS | _RATIO_FIELDS:
                tot = sum(non_zero) / len(non_zero) if non_zero else 0.0
            else:
                tot = sum(_safe(v) or 0.0 for v in annual_vals)
            ws.cell(row=row_num, column=COL_TOTAL,
                    value=_safe(tot)).number_format = _num_format(field)

    _apply_column_formatting(ws, n_years)
    ws.freeze_panes = "L7"


def _write_summary(ws, result: AssemblyResult) -> None:
    """Write the Summary sheet: project metadata + scalar KPIs."""
    ws.cell(row=1, column=COL_LABEL, value=result.project_name).font = Font(bold=True, size=14)

    ws.cell(row=3, column=COL_LABEL, value="Metric").font = _BOLD_FONT
    ws.cell(row=3, column=COL_CONSTANT, value="Value").font = _BOLD_FONT
    ws.cell(row=3, column=COL_UNIT, value="Units").font = _BOLD_FONT

    # Build KPI rows
    kpis: list[tuple[str, Any, str]] = []

    irr_out = result.outputs.get("IRR_001")
    if irr_out is not None:
        kpis.append(("Project IRR",            _safe(irr_out.project_irr),            "%"))
        kpis.append(("Equity IRR",             _safe(irr_out.equity_irr),             "%"))
        kpis.append(("Project NPV",            _safe(irr_out.project_npv),            "DKKk"))
        kpis.append(("Equity NPV",             _safe(irr_out.equity_npv),             "DKKk"))
        pb = irr_out.payback_period_months
        kpis.append(("Payback Period",         pb if pb >= 0 else "Never",            "months"))

    capex_out = result.outputs.get("CAPEX_001")
    if capex_out is not None:
        kpis.append(("Total CAPEX",            _safe(capex_out.cumulative_capex[-1]), "DKKk"))

    debt_out = result.outputs.get("DEBT_001")
    if debt_out is not None:
        kpis.append(("Min DSCR",               _safe(debt_out.min_dscr),             "x"))
        kpis.append(("Total Interest",         _safe(debt_out.total_interest),        "DKKk"))

    pl_out = result.outputs.get("PL_001")
    if pl_out is not None:
        kpis.append(("Lifetime Net Income",    _safe(sum(pl_out.net_income)),         "DKKk"))

    wacc_out = result.outputs.get("WACC_001")
    if wacc_out is not None:
        kpis.append(("WACC",                   _safe(wacc_out.wacc),                  "%"))
        kpis.append(("Blended CoE",            _safe(wacc_out.blended_cost_of_equity),"%"))

    _PCT_SUMMARY = {"Project IRR", "Equity IRR", "WACC", "Blended CoE"}
    _RATIO_SUMMARY = {"Min DSCR"}

    for i, (label, value, units) in enumerate(kpis):
        row = i + 4
        ws.cell(row=row, column=COL_LABEL, value=label)
        cell = ws.cell(row=row, column=COL_CONSTANT, value=value)
        ws.cell(row=row, column=COL_UNIT, value=units)
        if label in _PCT_SUMMARY and isinstance(value, float):
            cell.number_format = FMT_PCT
        elif label in _RATIO_SUMMARY and isinstance(value, float):
            cell.number_format = FMT_RATIO
        elif units == "DKKk" and isinstance(value, float):
            cell.number_format = FMT_DKKK


# ============================================================================
# CONFIG → MODULE_ID MAPPING
# ============================================================================

# Maps ProjectConfig field name → module_id used in outputs dict
_CONFIG_FIELD_TO_MODULE: dict[str, str] = {
    "wacc": "WACC_001",
    "price_curves": "PRICE_CURVES_001",
    "rev_pv": "REV_001",
    "rev_bess": "REV_002",
    "rev_wind": "REV_003",
    "opex_pv": "OPEX_001",
    "opex_bess": "OPEX_002",
    "opex_wind": "OPEX_003",
    "capex": "CAPEX_001",
    "bess_repow": "BESS_REPOW_001",
    "constr_finance": "CONSTR_FINANCE_001",
    "debt": "DEBT_001",
    "debt_sculpt": "DEBT_SCULPT_001",
    "debt_refi": "DEBT_REFI_001",
    "debt_linear": "DEBT_LINEAR_001",
    "shl": "SHL_001",
    "vat_facility": "VAT_FACILITY_001",
    "dsra": "DSRA_001",
    "ppa_cfd": "PPA_CFD_001",
    "repow_debt": "REPOW_DEBT_001",
    "bridge": "BRIDGE_FACILITY_001",
    "cash_sweep_module": "CASH_SWEEP_001",
    "mra": "MRA_001",
    "decom": "DECOM_PROVISION_001",
    "imbalance_fee": "IMBALANCE_FEE_001",
    "tax_lt": "TAX_LT_001",
    "tax_de": "TAX_DE_001",
    "tax_au": "TAX_AU_001",
    "working_capital": "WORKING_CAPITAL_001",
    "sources_uses": "SOURCES_USES_001",
    "valuation": "VALUATION_001",
    "breakeven": "BREAKEVEN_001",
    "model_checks": "MODEL_CHECKS_001",
    "dashboard": "DASHBOARD_001",
}

# Timeline fields injected by engine — skip on Assumptions sheet
_TIMELINE_FIELDS = {"periods", "start_year", "start_month"}

_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # pale yellow


def _sanitize_name(name: str) -> str:
    """Sanitize a string for use as an Excel named range."""
    return re.sub(r"[^A-Za-z0-9_]", "_", name)[:250]


# ============================================================================
# ASSUMPTIONS SHEET
# ============================================================================

def _write_assumptions_sheet(
    wb, ws, config: ProjectConfig,
) -> dict[str, str]:
    """Write scalar assumptions to an Assumptions sheet.

    Returns assumption_map: {asn_name: "Assumptions!$B${row}"}
    where asn_name = "asn_{module_id}_{field_name}".
    """
    ws.cell(row=1, column=1, value="Assumptions").font = Font(bold=True, size=14)
    ws.cell(row=3, column=1, value="Module").font = _BOLD_FONT
    ws.cell(row=3, column=2, value="Value").font = _BOLD_FONT
    ws.cell(row=3, column=3, value="Unit").font = _BOLD_FONT
    ws.cell(row=3, column=4, value="Field").font = _BOLD_FONT

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 28

    assumption_map: dict[str, str] = {}
    row = 4

    # StatementConfig scalars (opening balances)
    sc = config.statements
    for field_name in type(sc).model_fields:
        val = getattr(sc, field_name, None)
        if isinstance(val, (int, float, bool, str)) and field_name not in _TIMELINE_FIELDS:
            asn_name = _sanitize_name(f"asn_statements_{field_name}")
            ws.cell(row=row, column=1, value=f"Statements.{field_name}")
            cell = ws.cell(row=row, column=2, value=val)
            cell.fill = _INPUT_FILL
            ws.cell(row=row, column=4, value=field_name)
            ref = f"Assumptions!$B${row}"
            assumption_map[asn_name] = ref
            try:
                dn = DefinedName(asn_name, attr_text=f"'{ws.title}'!$B${row}")
                wb.defined_names.add(dn)
            except Exception:
                pass  # skip if name conflict
            row += 1

    # Module configs
    for cfg_field, mod_id in _CONFIG_FIELD_TO_MODULE.items():
        inp = getattr(config, cfg_field, None)
        if inp is None:
            continue
        if not hasattr(inp, "model_fields"):
            continue  # TaxConfig or other non-Pydantic
        for field_name in type(inp).model_fields:
            if field_name in _TIMELINE_FIELDS:
                continue
            val = getattr(inp, field_name, None)
            if isinstance(val, (int, float, bool)):
                asn_name = _sanitize_name(f"asn_{mod_id}_{field_name}")
                ws.cell(row=row, column=1, value=f"{mod_id}.{field_name}")
                cell = ws.cell(row=row, column=2, value=val)
                cell.fill = _INPUT_FILL
                ws.cell(row=row, column=4, value=field_name)
                ref = f"Assumptions!$B${row}"
                assumption_map[asn_name] = ref
                try:
                    dn = DefinedName(asn_name, attr_text=f"'{ws.title}'!$B${row}")
                    wb.defined_names.add(dn)
                except Exception:
                    pass
                row += 1

    # TaxConfig scalars
    if config.tax is not None:
        tc = config.tax
        for field_name in type(tc).model_fields:
            val = getattr(tc, field_name, None)
            if isinstance(val, (int, float, bool, str)) and field_name not in _TIMELINE_FIELDS:
                asn_name = _sanitize_name(f"asn_TAX_001_{field_name}")
                ws.cell(row=row, column=1, value=f"TAX_001.{field_name}")
                cell = ws.cell(row=row, column=2, value=val)
                cell.fill = _INPUT_FILL
                ws.cell(row=row, column=4, value=field_name)
                ref = f"Assumptions!$B${row}"
                assumption_map[asn_name] = ref
                try:
                    dn = DefinedName(asn_name, attr_text=f"'{ws.title}'!$B${row}")
                    wb.defined_names.add(dn)
                except Exception:
                    pass
                row += 1

    return assumption_map


# ============================================================================
# INPUTS SHEET
# ============================================================================

def _write_inputs_sheet(
    wb, ws, config: ProjectConfig, n: int, sy: int, sm: int,
) -> dict[tuple[str, str], int]:
    """Write time-series array inputs to an Inputs sheet.

    Returns input_row_map: {(module_id, field_name): row_number}
    """
    ws.cell(row=1, column=1, value="Inputs").font = Font(bold=True, size=14)

    # Row 2: period end dates
    for p in range(n):
        dt = _period_end_date(sy, sm, p)
        cell = ws.cell(row=2, column=COL_PERIOD_0 + p, value=dt)
        cell.number_format = "DD-MMM-YY"

    # Row 3: column headers
    ws.cell(row=3, column=1, value="Module.Field").font = _BOLD_FONT
    ws.cell(row=3, column=2, value="Unit").font = _BOLD_FONT
    for p in range(n):
        ws.cell(row=3, column=COL_PERIOD_0 + p, value=p + 1).font = _BOLD_FONT

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14

    input_row_map: dict[tuple[str, str], int] = {}
    row = 4

    for cfg_field, mod_id in _CONFIG_FIELD_TO_MODULE.items():
        inp = getattr(config, cfg_field, None)
        if inp is None:
            continue
        if not hasattr(inp, "model_fields"):
            continue
        for field_name in type(inp).model_fields:
            if field_name in _TIMELINE_FIELDS:
                continue
            val = getattr(inp, field_name, None)
            if not isinstance(val, list):
                continue
            if len(val) == 0:
                continue
            # Only write lists of primitive values (skip lists of Pydantic models)
            if not isinstance(val[0], (int, float, str, bool)):
                continue

            ws.cell(row=row, column=1, value=f"{mod_id}.{field_name}")
            ws.cell(row=row, column=2, value=get_units(field_name))
            input_row_map[(mod_id, field_name)] = row

            fmt = _num_format(field_name)
            for p, v in enumerate(val):
                cell = ws.cell(row=row, column=COL_PERIOD_0 + p, value=_safe(v))
                cell.number_format = fmt
                cell.fill = _INPUT_FILL
            row += 1

    # Set period column widths
    for p in range(n):
        ws.column_dimensions[get_column_letter(COL_PERIOD_0 + p)].width = 11.46

    ws.freeze_panes = f"{get_column_letter(COL_PERIOD_0)}4"
    return input_row_map


# ============================================================================
# FORMULA LAYER — PL_001, CF_001, BS_001
# ============================================================================

# Modules with formula support in this pass
_FORMULA_MODULES = {"PL_001", "CF_001", "BS_001"}


def _cell_ref(sheet: str, row: int, period: int) -> str:
    """Build an absolute cell reference like 'Revenue!L27' for a period column."""
    c = col_letter(period_col(period))
    # If same-sheet, caller should strip the sheet prefix
    return f"'{sheet}'!{c}{row}"


def _same_sheet_ref(row: int, period: int) -> str:
    """Build a same-sheet cell reference like 'L27'."""
    return f"{col_letter(period_col(period))}{row}"


def _same_sheet_prev(row: int, period: int) -> str:
    """Build a same-sheet reference to the previous period column."""
    return f"{col_letter(period_col(period - 1))}{row}"


def _sum_multi_rows(sheet: str, rows: list[int], period: int) -> str:
    """Build a SUM formula across multiple rows on the same sheet at a given period.

    Returns e.g. "'Revenue'!L27+'Revenue'!L58" for two revenue rows.
    If single row, returns a simple reference without SUM.
    """
    c = col_letter(period_col(period))
    refs = [f"'{sheet}'!{c}{r}" for r in rows]
    return "+".join(refs) if refs else "0"


def _find_upstream_rows(
    sheet: str, field_name: str, result: AssemblyResult,
) -> list[int]:
    """Find all ROW_MAP rows on `sheet` with `field_name` whose module has output."""
    rows = []
    for (sh, mid, f), row in ROW_MAP.items():
        if sh == sheet and f == field_name and mid in result.outputs:
            rows.append(row)
    return sorted(rows)


def _build_pl_refs(period: int, result: AssemblyResult) -> dict[str, str]:
    """Build refs dict for PL_001.get_excel_formulas() at a given period."""
    # gross_revenue = sum of all REV net_revenue rows on Revenue sheet
    rev_rows = _find_upstream_rows("Revenue", "net_revenue", result)
    opex_rows = _find_upstream_rows("Costs", "total_opex", result)

    # PL_001 rows on Statements sheet (FS_Monthly maps to "Statements" in ROW_MAP)
    ebitda_row = ROW_MAP[("Statements", "PL_001", "ebitda")]
    dep_row = ROW_MAP[("Statements", "PL_001", "depreciation")]
    ebit_row = ROW_MAP[("Statements", "PL_001", "ebit")]
    int_row = ROW_MAP[("Statements", "PL_001", "interest_expense")]
    ebt_row = ROW_MAP[("Statements", "PL_001", "ebt")]
    tax_row = ROW_MAP[("Statements", "PL_001", "tax_charge")]

    return {
        "gross_revenue":    _sum_multi_rows("Revenue", rev_rows, period) if rev_rows else "0",
        "total_opex":       _sum_multi_rows("Costs", opex_rows, period) if opex_rows else "0",
        "ebitda":           _same_sheet_ref(ebitda_row, period),
        "depreciation":     _same_sheet_ref(dep_row, period),
        "ebit":             _same_sheet_ref(ebit_row, period),
        "interest_expense": _same_sheet_ref(int_row, period),
        "ebt":              _same_sheet_ref(ebt_row, period),
        "tax_charge":       _same_sheet_ref(tax_row, period),
    }


def _build_cf_refs(
    period: int, result: AssemblyResult, assumption_map: dict[str, str],
) -> dict[str, str]:
    """Build refs dict for CF_001.get_excel_formulas() at a given period."""
    # Same-sheet references (FS_Monthly = Statements in ROW_MAP)
    ni_row = ROW_MAP[("Statements", "PL_001", "net_income")]
    dep_row = ROW_MAP[("Statements", "PL_001", "depreciation")]
    wc_row = ROW_MAP[("Statements", "CF_001", "working_capital_change")]

    # Cross-sheet CAPEX
    capex_row = ROW_MAP[("Costs", "CAPEX_001", "total_capex_monthly")]

    # Cross-sheet debt — sum all drawdown/principal/interest rows that have output
    draw_rows = _find_upstream_rows("Debt", "drawdown", result)
    princ_rows = _find_upstream_rows("Debt", "principal", result)
    # Also check total_principal for DEBT_LINEAR_001
    for (sh, mid, f), r in ROW_MAP.items():
        if sh == "Debt" and f == "total_principal" and mid in result.outputs:
            if r not in princ_rows:
                princ_rows.append(r)
    princ_rows.sort()
    int_rows = _find_upstream_rows("Debt", "interest", result)

    cfo_row = ROW_MAP[("Statements", "CF_001", "cfo")]
    cfi_row = ROW_MAP[("Statements", "CF_001", "cfi")]
    cff_row = ROW_MAP[("Statements", "CF_001", "cff")]
    div_row = ROW_MAP[("Statements", "CF_001", "dividends_paid")]
    ncf_row = ROW_MAP[("Statements", "CF_001", "net_cash_flow")]
    closing_row = ROW_MAP[("Statements", "CF_001", "closing_cash")]

    # Previous period closing cash; period 0 uses opening balance from Assumptions
    if period == 0:
        asn_key = "asn_statements_opening_cash_DKKk"
        prev_cash = assumption_map.get(asn_key, "0")
    else:
        prev_cash = _same_sheet_prev(closing_row, period)

    return {
        "net_income":       _same_sheet_ref(ni_row, period),
        "depreciation":     _same_sheet_ref(dep_row, period),
        "wc_change":        _same_sheet_ref(wc_row, period),
        "capex":            _cell_ref("Costs", capex_row, period),
        "debt_drawdown":    _sum_multi_rows("Debt", draw_rows, period) if draw_rows else "0",
        "principal":        _sum_multi_rows("Debt", princ_rows, period) if princ_rows else "0",
        "interest_paid":    _sum_multi_rows("Debt", int_rows, period) if int_rows else "0",
        "dividends":        _same_sheet_ref(div_row, period),
        "prev_closing_cash": prev_cash,
        "cfo":              _same_sheet_ref(cfo_row, period),
        "cfi":              _same_sheet_ref(cfi_row, period),
        "cff":              _same_sheet_ref(cff_row, period),
        "net_cash_flow":    _same_sheet_ref(ncf_row, period),
    }


def _build_bs_refs(
    period: int, result: AssemblyResult, assumption_map: dict[str, str],
) -> dict[str, str]:
    """Build refs dict for BS_001.get_excel_formulas() at a given period."""
    # Same-sheet row numbers (Statements)
    fa_gross_row = ROW_MAP[("Statements", "BS_001", "fixed_assets_gross")]
    acc_dep_row = ROW_MAP[("Statements", "BS_001", "accumulated_depreciation")]
    cash_row = ROW_MAP[("Statements", "BS_001", "cash")]
    ta_row = ROW_MAP[("Statements", "BS_001", "total_assets")]
    debt_row = ROW_MAP[("Statements", "BS_001", "debt_balance")]
    ce_row = ROW_MAP[("Statements", "BS_001", "contributed_equity")]
    re_row = ROW_MAP[("Statements", "BS_001", "retained_earnings")]
    te_row = ROW_MAP[("Statements", "BS_001", "total_equity")]
    ni_row = ROW_MAP[("Statements", "PL_001", "net_income")]
    div_row = ROW_MAP[("Statements", "CF_001", "dividends_paid")]
    tle_row = ROW_MAP[("Statements", "BS_001", "total_liabilities_equity")]

    # Cross-sheet CAPEX and depreciation
    capex_row = ROW_MAP[("Costs", "CAPEX_001", "total_capex_monthly")]
    dep_row = ROW_MAP[("Statements", "PL_001", "depreciation")]

    # Debt closing balance — sum all on Debt sheet
    debt_bal_rows = _find_upstream_rows("Debt", "closing_balance", result)

    # Previous period refs — period 0 uses 0 or Assumptions
    if period == 0:
        prev_fa = "0"
        prev_acc_dep = "0"
        asn_re = assumption_map.get("asn_statements_opening_retained_earnings_DKKk", "0")
        prev_retained = asn_re
    else:
        prev_fa = _same_sheet_prev(fa_gross_row, period)
        prev_acc_dep = _same_sheet_prev(acc_dep_row, period)
        prev_retained = _same_sheet_prev(re_row, period)

    asn_ce = assumption_map.get("asn_statements_opening_contributed_equity_DKKk", "0")

    return {
        "prev_fa_gross":            prev_fa,
        "capex":                    _cell_ref("Costs", capex_row, period),
        "prev_acc_dep":             prev_acc_dep,
        "depreciation":             _same_sheet_ref(dep_row, period),
        "fa_gross":                 _same_sheet_ref(fa_gross_row, period),
        "acc_dep":                  _same_sheet_ref(acc_dep_row, period),
        "cash":                     _same_sheet_ref(cash_row, period),
        "debt":                     _sum_multi_rows("Debt", debt_bal_rows, period) if debt_bal_rows else "0",
        "contributed_equity":       asn_ce,
        "prev_retained":            prev_retained,
        "net_income":               _same_sheet_ref(ni_row, period),
        "dividends":                _same_sheet_ref(div_row, period),
        "total_equity":             _same_sheet_ref(te_row, period),
        "total_assets":             _same_sheet_ref(ta_row, period),
        "total_liabilities_equity": _same_sheet_ref(tle_row, period),
    }


def _try_get_formula(
    module_id: str, field_name: str, period: int,
    result: AssemblyResult, assumption_map: dict[str, str],
) -> str | None:
    """Try to get an Excel formula for a field. Returns None if not available.

    Layer 1: Call get_excel_formulas() for supported modules.
    Layer 2: Return None → caller writes computed value.
    """
    if module_id not in _FORMULA_MODULES:
        return None

    try:
        if module_id == "PL_001":
            refs = _build_pl_refs(period, result)
            formulas = PL_001.get_excel_formulas(refs)
        elif module_id == "CF_001":
            refs = _build_cf_refs(period, result, assumption_map)
            formulas = CF_001.get_excel_formulas(refs)
        elif module_id == "BS_001":
            refs = _build_bs_refs(period, result, assumption_map)
            formulas = BS_001.get_excel_formulas(refs)
        else:
            return None
    except (KeyError, TypeError):
        return None

    return formulas.get(field_name)


# ============================================================================
# PUBLIC API
# ============================================================================

def write_workbook(
    result: AssemblyResult,
    config: ProjectConfig,
    output_path: str,
) -> None:
    """
    Write result to a .xlsx workbook at output_path.

    Sheets: Assumptions, Inputs, Cover, Revenue, Costs, Debt,
            FS_Monthly, FS_Annual, Summary.

    Two-layer strategy: PL_001/CF_001/BS_001 write Excel formulas where
    available; all other modules write Python-computed values (fallback).
    """
    # Set currency label from config before writing any cells
    from assembly.cell_mapper import set_currency_label
    set_currency_label(config.timeline.currency_label)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    n  = result.periods
    sy = result.start_year
    sm = result.start_month

    # Step 1: Assumptions sheet (scalar inputs with named ranges)
    ws_asn = wb.create_sheet("Assumptions")
    assumption_map = _write_assumptions_sheet(wb, ws_asn, config)

    # Step 2: Inputs sheet (time-series array inputs)
    ws_inp = wb.create_sheet("Inputs")
    _write_inputs_sheet(wb, ws_inp, config, n, sy, sm)

    # Step 3: Create calculation sheets upfront in display order
    for name in SHEETS:
        wb.create_sheet(name)

    capex_out = result.outputs.get("CAPEX_001")
    capex_monthly = capex_out.total_capex_monthly if capex_out else None

    # Step 4: Populate calculation sheets — formulas on FS_Monthly, values elsewhere
    for sheet_name in ["Revenue", "Costs", "Debt", "FS_Monthly"]:
        ws = wb[sheet_name]
        _write_header_rows(ws, sheet_name, n, sy, sm, capex_monthly=capex_monthly)
        _write_time_series_rows(ws, result, assumption_map=assumption_map)
        if sheet_name == "Debt":
            _write_wacc_scalars(ws, result)
        _write_subsection_labels(ws, sheet_name)
        _apply_column_formatting(ws, n)
        ws.freeze_panes = "L7"

    _write_fs_annual(wb["FS_Annual"], result, config)
    _write_subsection_labels(wb["FS_Annual"], "FS_Annual")
    _write_summary(wb["Summary"], result)

    # Step 5: Write Cover LAST (needs Summary to be populated for KPI refs)
    write_cover(wb["Cover"], result, config, wb)

    # Step 6: Apply formatting and save
    apply_formatting(wb, result)
    wb.save(output_path)
