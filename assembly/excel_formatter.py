"""
assembly/excel_formatter.py

Visual formatting layer for the Excel workbook.
Applied AFTER all data is written. Never changes cell values (except section
header labels written into empty rows).

Two style modes:
  FormatStyle.F1F9       — clean F1F9 renewables modelling standard (default)
  FormatStyle.EE_LEGACY  — EE/PwC fill-based section coloring

CREATED:  2026-03-17
MODIFIED: 2026-03-30
"""

from __future__ import annotations

from enum import Enum

from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from assembly.cell_mapper import (
    ROW_MAP, COL_LABEL, COL_UNIT, COL_TOTAL, COL_PERIOD_0, COL_SPACER,
)


# ============================================================================
# FORMAT STYLE CONFIG
# ============================================================================

class FormatStyle(str, Enum):
    F1F9 = "f1f9"          # clean, minimal, professional (default)
    EE_LEGACY = "ee"       # EE/PwC teal/grey section fills

# Module-level default — can be overridden by caller
DEFAULT_STYLE = FormatStyle.F1F9

# ============================================================================
# F1F9 COLOR CONSTANTS (from F1F9-Example-Renewables-Model-01a.xlsm)
# ============================================================================

# Font colors — F1F9 standard
FONT_DEFAULT        = "000000"  # black  — same-sheet formulas
FONT_INPUT          = "0000FF"  # blue   — hardcoded inputs
FONT_IMPORT         = "FF0000"  # red    — cross-sheet references
FONT_EXPORT         = "FF0000"  # red    — output/export rows
FONT_WHITE          = "FFFFFF"  # white  — on dark fills
FONT_GREEN          = "00B050"  # green  — check pass indicator

# Structure fills — F1F9 standard
FILL_SECTION_HEADER = "DDDDDD"  # light grey — section header band
FILL_SUBSECTION     = "EAEAEA"  # medium-light grey — sub-section band
FILL_INPUT_BG       = "F8F8F8"  # near-white — input cell subtle background
FILL_ATTENTION      = "FFFF99"  # pale yellow — key assumption cells
FILL_CHECK_PASS     = "99FF66"  # green — check OK
FILL_CHECK_FAIL     = "FF0000"  # red — check fail
FILL_ALT_HEADER     = "D9D9D9"  # alternate grey
FILL_OPTIMISE       = "CCFFFF"  # cyan — optimisation parameter cells
FILL_OUTPUT         = "99CCFF"  # pale blue — output summary cells
FILL_PHASE_STRIPE   = "F8F8F8"  # near-white (same as input bg)
FILL_TOTAL_COL      = "F2F2F2"  # light grey — col J total/avg

# EE legacy fills (kept for backward compat)
FILL_COL_HEADER     = "44546A"  # slate blue — EE column header row
FILL_SUBSECTION_LABEL = "A6A6A6"  # mid grey — EE reference model
FILL_COVER_HEADER   = "008080"  # teal
FILL_COVER_KPI      = "EBF3FB"
FILL_COVER_KEY_ROW  = "F2F2F2"
FILL_COUNTER_FLOW   = "DDDDDD"
FILL_INPUT          = "FFF2CC"  # EE legacy pale yellow — editable cells

# Tab colors — F1F9 convention (from the spec)
TAB_COLORS_F1F9 = {
    "Cover":      "FFFFFF",  # white
    "Input":      "FFFF99",  # pale yellow
    "Revenue":    None,      # default (no color) — calc sheets
    "Costs":      None,
    "Debt":       None,
    "FS_Monthly": "99CCFF",  # pale blue — output
    "FS_Annual":  "99CCFF",
    "Summary":    "CCFF99",  # pale green — checks/validation
}

# Tab colors — EE legacy
TAB_COLORS_EE = {
    "Cover":      "008080",
    "Revenue":    "28837D",
    "Costs":      "28837D",
    "Debt":       "28837D",
    "FS_Monthly": "44546A",
    "FS_Annual":  "44546A",
    "Summary":    "1F4E79",
}

TAB_COLORS = TAB_COLORS_EE


# ============================================================================
# FONT CONSTANTS
# ============================================================================

# F1F9 uses Arial; EE_LEGACY uses Calibri
FONT_NAME_F1F9 = "Arial"
FONT_NAME_EE = "Calibri"

def _font_name(style: FormatStyle = None) -> str:
    s = style or DEFAULT_STYLE
    return FONT_NAME_F1F9 if s == FormatStyle.F1F9 else FONT_NAME_EE


# ============================================================================
# NUMBER FORMAT CONSTANTS
# ============================================================================

# F1F9 number formats (4-section: positive_);(negative);dash;text)
FMT_DKKK_F1F9 = '#,##0_);\\(#,##0\\);"-  ";" "@" "'
FMT_DATE_F1F9 = 'dd mmm yy_);\\(###0\\);"-  ";" "@" "'
FMT_PCT_F1F9  = '0.00_)%_);\\(0.00\\)%_);"-  ";" "@" "'
FMT_RATE_4DP  = '#,##0.0000_);\\(#,##0.0000\\);"-  ";" "@" "'
FMT_1DP_F1F9  = '#,##0.0_);\\(#,##0.0\\);"-  ";" "@" "'
FMT_2DP_F1F9  = '#,##0.00_);\\(#,##0.00\\);"-  ";" "@" "'

# EE legacy formats
FMT_DKKK_EE   = '#,##0'

# Active defaults (used by _get_field_format)
FMT_DKKK    = '#,##0'          # overridden per style in apply
FMT_PCT     = '0.0%'           # percentages
FMT_RATIO   = '0.00"x"'        # DSCR, coverage ratios
FMT_FACTOR  = '0.000'          # indexation factors, betas
FMT_INTEGER = '#,##0'          # MWh, periods, years
FMT_DATE    = 'DD-MMM-YY'      # period end dates (row 2)
FMT_YEAR    = '0'              # year integers
FMT_GENERAL = 'General'

# Field name -> format string
FIELD_FORMATS: dict[str, str] = {
    # Percentages
    "wacc": FMT_PCT, "blended_cost_of_equity": FMT_PCT,
    "ppa_cost_of_equity": FMT_PCT, "merchant_cost_of_equity": FMT_PCT,
    "cost_of_debt_posttax": FMT_PCT, "ppa_erp": FMT_PCT,
    "merchant_erp": FMT_PCT, "project_irr": FMT_PCT, "equity_irr": FMT_PCT,
    "ebitda_margin": FMT_PCT, "tax_rate": FMT_PCT,
    "price_indexation_factor": FMT_FACTOR,
    "discharge_indexation_factor": FMT_FACTOR,
    "charge_indexation_factor": FMT_FACTOR,
    # Ratios
    "dscr_annual": FMT_RATIO, "llcr": FMT_RATIO,
    "dscr_achieved": FMT_RATIO, "dscr_covenant_applicable": FMT_RATIO,
    "llcr_series": FMT_RATIO,
    "levered_beta": FMT_FACTOR, "debt_to_equity_ratio": FMT_FACTOR,
    # Integers
    "net_production_MWh": FMT_INTEGER, "discharge_volume": FMT_INTEGER,
    "goo_volume": FMT_INTEGER, "payback_period_months": FMT_INTEGER,
    # Everything else defaults to FMT_DKKK
}


def _get_field_format(field_name: str) -> str:
    """Return number format string for a field. Defaults to FMT_DKKK."""
    return FIELD_FORMATS.get(field_name, FMT_DKKK)


# ============================================================================
# ROW CLASSIFICATION SETS
# ============================================================================

# Red font — values used by other sheets
EXPORT_FIELDS = {
    "net_revenue", "total_ppa_revenue", "merchant_revenue", "goo_revenue",
    "total_opex", "total_capex_monthly", "cumulative_capex",
    "interest", "principal_repayment", "closing_balance", "dscr_annual",
    "tax_charge_accrued", "tax_paid", "tax_depreciation",
    "net_income", "cfo", "cfi", "cff", "net_cash_flow", "closing_cash",
    "total_assets", "total_liabilities_equity",
}

# Blue font — values pulled from other sheets
IMPORT_FIELDS = {
    "gross_revenue", "depreciation", "interest_expense",
    "capex_monthly", "debt_drawdown",
    "net_production_MWh", "discharge_volume",
}

# Thin top+bottom border — subtotals
SUBTOTAL_FIELDS = {
    "ebitda", "ebit", "ebt", "cfo", "cfi", "cff",
    "net_cash_flow", "total_assets", "total_liabilities_equity", "net_revenue",
}

# Bold + medium top border — totals
TOTAL_FIELDS = {
    "net_income", "closing_cash", "project_irr", "equity_irr", "cumulative_capex",
}

# Section header rows: {sheet_name: {row_number: label}}
# Row numbers chosen to sit in empty gaps between module groups in ROW_MAP
SECTION_HEADER_ROWS = {
    "Revenue":    {29: "BESS Revenue", 60: "Wind Revenue", 83: "PPA / CfD"},
    "Costs":      {19: "BESS OPEX", 26: "Capital Expenditure",
                   36: "Wind OPEX", 43: "BESS Repowering"},
    "Debt":       {21: "WACC", 32: "Shareholder Loan",
                   38: "VAT Facility", 45: "Sculpted Debt"},
    "FS_Monthly": {15: "Cash Flow", 33: "Balance Sheet"},
    "FS_Annual":  {15: "Cash Flow", 33: "Balance Sheet"},
}

# Row height spec (points) — F1F9 standard
ROW_HEIGHTS_F1F9 = {
    "banner":         25.15,  # row 1 only
    "sub_banner":     13.05,  # rows 2-4 header band
    "year_row":       13.05,
    "col_header":     13.05,
    "spacer":          5.25,  # thin spacer rows
    "section_header": 13.05,  # same as data (F1F9 uses grey fill, not height)
    "data":           13.05,  # ALL data rows
    "subtotal":       13.05,
    "total":          13.05,
}

# Row height spec — EE legacy
ROW_HEIGHTS_EE = {
    "banner":         22,
    "sub_banner":     18,
    "year_row":       15,
    "col_header":     20,
    "spacer":          5,
    "section_header": 18,
    "data":           15,
    "subtotal":       16,
    "total":          17,
}

ROW_HEIGHTS = ROW_HEIGHTS_EE


# ============================================================================
# HELPERS
# ============================================================================

def _get_sheet_rows(sheet_name: str) -> list[tuple[int, str, str]]:
    """Return [(row_number, module_id, field_name)] for a given sheet."""
    logical = "Statements" if sheet_name in ("FS_Monthly", "FS_Annual") else sheet_name
    rows = []
    for (sht, mod, field), row in ROW_MAP.items():
        if sht == logical:
            rows.append((row, mod, field))
    return sorted(rows)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color: str = FONT_DEFAULT, bold: bool = False,
          size: int = 10, style: FormatStyle = None) -> Font:
    return Font(color=hex_color, bold=bold, name=_font_name(style), size=size)


def _border_subtotal() -> Border:
    """Thin top and bottom border for subtotal rows (EE legacy)."""
    thin = Side(border_style="thin")
    return Border(top=thin, bottom=thin)


def _border_total() -> Border:
    """Medium top border for total rows (EE legacy)."""
    return Border(top=Side(border_style="medium"))


# F1F9 borders (from spec)
_SIDE_GREY = Side(style='thin', color='808080')
_SIDE_BLACK_THIN = Side(style='thin', color='000000')
_SIDE_BLACK_MED = Side(style='medium', color='000000')

def _border_f1f9_row_top() -> Border:
    """F1F9: thin grey top border for standard row separator."""
    return Border(top=_SIDE_GREY)

def _border_f1f9_subtotal() -> Border:
    """F1F9: thin black bottom border for subtotals."""
    return Border(bottom=_SIDE_BLACK_THIN)

def _border_f1f9_total() -> Border:
    """F1F9: medium black bottom border for grand totals."""
    return Border(bottom=_SIDE_BLACK_MED)


# ============================================================================
# FORMAT FUNCTIONS (Prompt 2)
# ============================================================================

def _format_header_rows(ws, n_cols: int, style: FormatStyle = None) -> None:
    """Format header band rows."""
    s = style or DEFAULT_STYLE
    last_col = COL_PERIOD_0 + n_cols - 1

    if s == FormatStyle.EE_LEGACY:
        # Row 5: col header — slate fill, white bold
        for c in range(1, last_col + 1):
            cell = ws.cell(row=5, column=c)
            cell.fill = _fill(FILL_COL_HEADER)
            cell.font = _font(FONT_WHITE, bold=True)
        # Row 3: phase stripe — near-white fill
        for c in range(COL_PERIOD_0, last_col + 1):
            ws.cell(row=3, column=c).fill = _fill(FILL_PHASE_STRIPE)
    else:
        # F1F9: row 1 title (14pt), rows 2-4 bold headers, no fills
        ws.cell(row=1, column=COL_LABEL).font = _font(bold=False, size=14, style=s)
        for r in (2, 3, 4):
            ws.cell(row=r, column=COL_LABEL).font = _font(bold=True, style=s)
        # Phase stripe: near-white fill
        for c in range(COL_PERIOD_0, last_col + 1):
            ws.cell(row=3, column=c).fill = _fill(FILL_PHASE_STRIPE)


def _format_data_rows(ws, sheet_name: str, n_cols: int, style: FormatStyle = None) -> None:
    """Apply font colors, borders, and col J fill to all data rows."""
    s = style or DEFAULT_STYLE
    last_data_col = COL_PERIOD_0 + n_cols - 1
    rows = _get_sheet_rows(sheet_name)

    for row_num, mod_id, field in rows:
        # Col J: total column fill
        ws.cell(row=row_num, column=COL_TOTAL).fill = _fill(FILL_TOTAL_COL)

        # Font color: F1F9 uses red for cross-sheet, blue for inputs
        if field in EXPORT_FIELDS:
            font_color = FONT_EXPORT
        elif field in IMPORT_FIELDS:
            font_color = FONT_IMPORT
        else:
            font_color = FONT_DEFAULT

        for c in range(COL_LABEL, last_data_col + 1):
            cell = ws.cell(row=row_num, column=c)
            existing_bold = cell.font.bold if cell.font else False
            cell.font = _font(font_color, bold=existing_bold, style=s)

        # Borders differ by style
        if s == FormatStyle.F1F9:
            if field in SUBTOTAL_FIELDS:
                for c in range(COL_LABEL, last_data_col + 1):
                    ws.cell(row=row_num, column=c).border = _border_f1f9_subtotal()
            elif field in TOTAL_FIELDS:
                for c in range(COL_LABEL, last_data_col + 1):
                    ws.cell(row=row_num, column=c).border = _border_f1f9_total()
                ws.cell(row=row_num, column=COL_LABEL).font = _font(font_color, bold=True, style=s)
        else:
            if field in SUBTOTAL_FIELDS:
                for c in range(COL_LABEL, last_data_col + 1):
                    ws.cell(row=row_num, column=c).border = _border_subtotal()
            elif field in TOTAL_FIELDS:
                for c in range(COL_LABEL, last_data_col + 1):
                    ws.cell(row=row_num, column=c).border = _border_total()
                ws.cell(row=row_num, column=COL_LABEL).font = _font(font_color, bold=True, style=s)


def _format_section_headers(ws, sheet_name: str, style: FormatStyle = None) -> None:
    """Write section header labels into empty rows."""
    s = style or DEFAULT_STYLE
    headers = SECTION_HEADER_ROWS.get(sheet_name, {})
    for row_num, lbl in headers.items():
        existing = ws.cell(row=row_num, column=COL_LABEL).value
        if existing is not None:
            continue
        if s == FormatStyle.F1F9:
            # F1F9: grey fill band, bold black text
            for c in range(COL_LABEL, COL_UNIT + 1):
                ws.cell(row=row_num, column=c).fill = _fill(FILL_SECTION_HEADER)
                ws.cell(row=row_num, column=c).font = _font(bold=True, style=s)
        else:
            # EE: teal fill, white text
            for c in range(COL_LABEL, COL_UNIT + 1):
                ws.cell(row=row_num, column=c).fill = _fill("28837D")
                ws.cell(row=row_num, column=c).font = _font(FONT_WHITE, bold=True, style=s)
        ws.cell(row=row_num, column=COL_LABEL).value = lbl


def _format_subsection_labels(ws, sheet_name: str, style: FormatStyle = None) -> None:
    """Apply sub-section label formatting."""
    s = style or DEFAULT_STYLE
    from assembly.cell_mapper import SUBSECTION_LABELS

    for (sht, row_num), lbl in SUBSECTION_LABELS.items():
        if sht != sheet_name:
            continue
        if s == FormatStyle.F1F9:
            # F1F9: lighter grey band, bold black
            for c in range(1, COL_UNIT + 1):
                ws.cell(row=row_num, column=c).fill = _fill(FILL_SUBSECTION)
                ws.cell(row=row_num, column=c).font = _font(bold=True, style=s)
            ws.row_dimensions[row_num].height = 13.05
        else:
            # EE: dark grey fill, white bold
            for c in range(1, COL_UNIT + 1):
                ws.cell(row=row_num, column=c).fill = _fill(FILL_SUBSECTION_LABEL)
                ws.cell(row=row_num, column=c).font = _font(FONT_WHITE, bold=True, style=s)
            ws.row_dimensions[row_num].height = 16


# ============================================================================
# LAYOUT FUNCTIONS (Prompt 3)
# ============================================================================

def _apply_number_formats(ws, sheet_name: str, n_cols: int) -> None:
    """Apply canonical number formats to all time series and total cells.

    Iterates every row in ROW_MAP for this sheet and applies the correct
    format to col J (total) and cols L+ (time series). The writer may have
    already set formats — this pass overwrites to ensure consistency.

    Row 2 (period end dates): FMT_DATE applied to all time series cols.
    Row 4 (year integers):    FMT_YEAR applied to all time series cols.
    """
    logical = "Statements" if sheet_name in ("FS_Monthly", "FS_Annual") else sheet_name

    # Header row formats
    date_fmt = FMT_DATE if sheet_name != "FS_Annual" else FMT_YEAR
    for col in range(COL_PERIOD_0, COL_PERIOD_0 + n_cols):
        ws.cell(row=2, column=col).number_format = date_fmt
        ws.cell(row=4, column=col).number_format = FMT_YEAR

    # Data row formats
    for (sht, mod, field), row_num in ROW_MAP.items():
        if sht != logical:
            continue
        fmt = _get_field_format(field)
        ws.cell(row=row_num, column=COL_TOTAL).number_format = fmt
        for p in range(n_cols):
            ws.cell(row=row_num, column=COL_PERIOD_0 + p).number_format = fmt


def _apply_row_grouping(ws, sheet_name: str) -> None:
    """Apply outline grouping to detail rows within each module section.

    Groups detail rows (outline_level=1, hidden=True) between the first
    and last row of each module section. Section header and total rows
    remain at outline_level=0 and are always visible.

    Sections with fewer than 4 rows are not grouped.
    """
    from assembly.cell_mapper import SUBSECTION_LABELS

    logical = "Statements" if sheet_name in ("FS_Monthly", "FS_Annual") else sheet_name
    module_rows: dict[str, list[int]] = {}
    for (sht, mod, field), row in ROW_MAP.items():
        if sht == logical:
            module_rows.setdefault(mod, []).append(row)

    # Subsection label rows must never be grouped
    label_rows = {row for (sht, row) in SUBSECTION_LABELS if sht == sheet_name}

    for mod_id, rows in module_rows.items():
        rows_sorted = sorted(rows)
        if len(rows_sorted) < 4:
            continue  # too few rows to group meaningfully

        detail_rows = [r for r in rows_sorted[1:-1] if r not in label_rows]
        for r in detail_rows:
            ws.row_dimensions[r].outline_level = 1
            ws.row_dimensions[r].hidden = True

    ws.sheet_view.showOutlineSymbols = True


def _apply_row_heights(ws, sheet_name: str) -> None:
    """Apply row heights by row type.

    Header rows 1-5: fixed heights.
    Data rows: height by field classification.
    Section header rows: section_header height.
    Blank spacer rows between sections: spacer height (5pt).
    """
    logical = "Statements" if sheet_name in ("FS_Monthly", "FS_Annual") else sheet_name

    # Build set of all row numbers that have data
    data_row_map: dict[int, str] = {}  # row -> field_name
    for (sht, mod, field), row in ROW_MAP.items():
        if sht == logical:
            data_row_map[row] = field

    if not data_row_map:
        return

    last_data_row = max(data_row_map.keys())

    # Header rows 1-5
    ws.row_dimensions[1].height = ROW_HEIGHTS["banner"]
    ws.row_dimensions[2].height = ROW_HEIGHTS["banner"]
    ws.row_dimensions[3].height = ROW_HEIGHTS["sub_banner"]
    ws.row_dimensions[4].height = ROW_HEIGHTS["year_row"]
    ws.row_dimensions[5].height = ROW_HEIGHTS["col_header"]

    # Section header rows
    section_rows = set(SECTION_HEADER_ROWS.get(sheet_name, {}).keys())
    for r in section_rows:
        ws.row_dimensions[r].height = ROW_HEIGHTS["section_header"]

    # Data rows
    for row_num, field in data_row_map.items():
        if field in SUBTOTAL_FIELDS:
            ws.row_dimensions[row_num].height = ROW_HEIGHTS["subtotal"]
        elif field in TOTAL_FIELDS:
            ws.row_dimensions[row_num].height = ROW_HEIGHTS["total"]
        else:
            ws.row_dimensions[row_num].height = ROW_HEIGHTS["data"]

    # Blank spacer rows between row 6 and last data row
    occupied = data_row_map.keys() | section_rows | {1, 2, 3, 4, 5}
    for r in range(6, last_data_row + 1):
        if r not in occupied:
            ws.row_dimensions[r].height = ROW_HEIGHTS["spacer"]


def _apply_print_setup(ws) -> None:
    """Configure print settings for bank-grade output.

    Landscape A3, fit to 1 page wide, repeat header rows and label
    columns on every printed page.
    """
    last_frozen_col = get_column_letter(COL_SPACER)  # K

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 8        # A3
    ws.page_setup.fitToWidth  = 1        # 1 page wide
    ws.page_setup.fitToHeight = 0        # unlimited pages tall
    ws.page_setup.scale       = 100      # ignored when fitToPage=True
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Repeat rows 1-6 and cols A-K on every printed page
    ws.print_title_rows = "1:6"
    ws.print_title_cols = f"A:{last_frozen_col}"

    # Margins in inches
    ws.page_margins = PageMargins(
        left=0.5, right=0.5,
        top=0.75, bottom=0.75,
        header=0.3, footer=0.3,
    )


# ============================================================================
# PUBLIC API
# ============================================================================

def apply_formatting(wb, result, style: FormatStyle = None) -> None:
    """Apply format standard to a fully written workbook.

    Args:
        wb: openpyxl Workbook
        result: AssemblyResult with periods/start_year/start_month
        style: FormatStyle.F1F9 (default) or FormatStyle.EE_LEGACY

    Call order is strict — do not reorder.
    """
    s = style or DEFAULT_STYLE
    global ROW_HEIGHTS
    ROW_HEIGHTS = ROW_HEIGHTS_F1F9 if s == FormatStyle.F1F9 else ROW_HEIGHTS_EE

    CALC_SHEETS = ["Revenue", "Costs", "Debt", "FS_Monthly", "FS_Annual"]
    n = result.periods

    for sheet_name in CALC_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        if sheet_name == "FS_Annual":
            years = set()
            for p in range(n):
                years.add(result.start_year + ((result.start_month - 1) + p) // 12)
            n_cols = len(years)
        else:
            n_cols = n

        _apply_number_formats(ws, sheet_name, n_cols)
        _format_header_rows(ws, n_cols, s)
        _format_data_rows(ws, sheet_name, n_cols, s)
        _format_section_headers(ws, sheet_name, s)
        _format_subsection_labels(ws, sheet_name, s)
        _apply_row_grouping(ws, sheet_name)
        _apply_row_heights(ws, sheet_name)
        _apply_print_setup(ws)

    for sheet_name in ["Cover", "Summary"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    _apply_tab_colors(wb, s)

    # Add Style Guide sheet in F1F9 mode
    if s == FormatStyle.F1F9:
        _add_style_guide(wb, s)


def _apply_tab_colors(wb, style: FormatStyle = None) -> None:
    """Apply tab colors by sheet category and style."""
    s = style or DEFAULT_STYLE
    colors = TAB_COLORS_F1F9 if s == FormatStyle.F1F9 else TAB_COLORS_EE
    for sheet_name, color in colors.items():
        if sheet_name in wb.sheetnames and color:
            wb[sheet_name].sheet_properties.tabColor = color


def _add_style_guide(wb, style: FormatStyle) -> None:
    """Add a Style Guide sheet demonstrating the formatting standard."""
    ws = wb.create_sheet("Style Guide")
    fname = _font_name(style)

    ws.column_dimensions['E'].width = 40.5
    ws.column_dimensions['F'].width = 12.5
    ws.column_dimensions['L'].width = 14
    ws.row_dimensions[1].height = 25.15

    ws.cell(row=2, column=5, value="Style Guide -- F1F9 Layout Standard").font = Font(
        name=fname, bold=True, size=10)

    # Font Colors section
    ws.cell(row=4, column=5, value="FONT COLORS").font = Font(name=fname, bold=True)
    ws.cell(row=5, column=5, value="Input value (hardcoded)").font = Font(name=fname, color="0000FF")
    ws.cell(row=5, column=12, value=100).font = Font(name=fname, color="0000FF")
    ws.cell(row=6, column=5, value="Cross-sheet reference").font = Font(name=fname, color="0000FF")
    ws.cell(row=6, column=12, value="=Rev!L11").font = Font(name=fname, color="0000FF")
    ws.cell(row=7, column=5, value="Within-sheet calculation").font = Font(name=fname, color="000000")
    ws.cell(row=7, column=12, value="=L5+L6").font = Font(name=fname)
    ws.cell(row=8, column=5, value="Output / export row").font = Font(name=fname, color="FF0000")
    ws.cell(row=8, column=12, value="=SUM(L5:L7)").font = Font(name=fname, color="FF0000")

    # Row Types section
    ws.cell(row=10, column=5, value="ROW TYPES").font = Font(name=fname, bold=True)
    ws.cell(row=11, column=5, value="Section header").font = Font(name=fname, bold=True)
    ws.cell(row=12, column=5, value="Data row").font = Font(name=fname)
    ws.cell(row=13, column=5, value="Data row").font = Font(name=fname)
    thin = Side(border_style="thin")
    ws.cell(row=14, column=5, value="Subtotal").font = Font(name=fname, bold=True)
    ws.cell(row=14, column=5).border = Border(top=thin)
    ws.cell(row=14, column=12).border = Border(top=thin)
    ws.row_dimensions[15].height = 5.25
    ws.cell(row=16, column=5, value="Grand total").font = Font(name=fname, bold=True)
    ws.cell(row=16, column=5).border = Border(top=thin, bottom=thin)
    ws.cell(row=16, column=12).border = Border(top=thin, bottom=thin)

    # Number Formats section
    f1f9_fmt = '#,##0_);(#,##0);"-  ";" "@" "'
    ws.cell(row=18, column=5, value="NUMBER FORMATS").font = Font(name=fname, bold=True)
    ws.cell(row=19, column=5, value="Positive kEUR").font = Font(name=fname)
    ws.cell(row=19, column=12, value=1234).number_format = f1f9_fmt
    ws.cell(row=20, column=5, value="Negative kEUR").font = Font(name=fname)
    ws.cell(row=20, column=12, value=-1234).number_format = f1f9_fmt
    ws.cell(row=21, column=5, value="Zero").font = Font(name=fname)
    ws.cell(row=21, column=12, value=0).number_format = f1f9_fmt
    ws.cell(row=22, column=5, value="Percentage").font = Font(name=fname)
    ws.cell(row=22, column=12, value=0.05).number_format = '0.0%'
    ws.cell(row=23, column=5, value="Ratio").font = Font(name=fname)
    ws.cell(row=23, column=12, value=1.35).number_format = '0.00"x"'
    ws.cell(row=24, column=5, value="Date").font = Font(name=fname)
    from datetime import date
    ws.cell(row=24, column=12, value=date(2026, 3, 1)).number_format = 'MMM-YY'

    # Column Layout section
    ws.cell(row=26, column=5, value="COLUMN LAYOUT").font = Font(name=fname, bold=True)
    for i, desc in enumerate([
        "A-D: Spacers (1.33 wide)",
        "E: Labels (40.5 wide)",
        "F: Constants (12.5 wide)",
        "G: Units (14.5 wide)",
        "H: Notes (hidden)",
        "L+: Time series data (11 wide)",
    ]):
        ws.cell(row=27+i, column=5, value=desc).font = Font(name=fname)
