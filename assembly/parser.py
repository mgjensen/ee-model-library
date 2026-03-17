"""
assembly/parser.py

Parses operator-uploaded .xlsx files into a flat dict of assumption values.

Two strategies are tried in order:
    1. Named ranges  — workbook-level defined names; each maps to a single cell.
    2. Assumptions sheet — sheet named "Assumptions" with col A = name, col B = value.

Also contains parse_vendor_model() for full EE vendor model format
(1.1 Input_NTD, 1.2 Input_M, 1.3 Input_A, 1.4 Global tabs).

Returns a flat dict {str: any} with string keys and numeric/string values.

Limitations:
    .xlsb files are NOT supported (openpyxl cannot read them).
    A ValueError is raised with instructions to save as .xlsx first.
"""

from __future__ import annotations

import os


def parse_xlsx(file_path: str) -> dict:
    """
    Parse assumption values from an Excel file.

    Parameters
    ----------
    file_path : Path to the .xlsx file.

    Returns
    -------
    dict
        Flat dict of {name: value} pairs. Keys are stripped of whitespace.
        Values are whatever openpyxl reads from the cell (float, int, str, bool, None).

    Raises
    ------
    FileNotFoundError
        If file_path does not exist.
    ValueError
        If the file extension is .xlsb (binary format not supported).
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path!r}")

    _, ext = os.path.splitext(file_path)
    if ext.lower() == ".xlsb":
        raise ValueError(
            f"Binary Excel files (.xlsb) are not supported. "
            f"Please open {os.path.basename(file_path)!r} in Excel, "
            f"choose 'Save As', and select 'Excel Workbook (.xlsx)'."
        )

    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Strategy 1: named ranges
    result = _read_named_ranges(wb)

    # Strategy 2: "Assumptions" sheet (merge, Strategy 2 wins on conflict)
    assumptions_data = _read_assumptions_sheet(wb)
    result.update(assumptions_data)

    return result


def _read_named_ranges(wb) -> dict:
    """
    Extract workbook-level defined names that resolve to a single cell.

    Only single-cell destinations are extracted (multi-cell ranges are skipped).
    """
    data: dict = {}

    for name in wb.defined_names:
        # openpyxl 3.1+: iterating yields string keys; look up the DefinedName object
        defined_name = wb.defined_names[name]
        # Each DefinedName may have multiple destinations (sheet, coord pairs)
        destinations = list(defined_name.destinations)
        if len(destinations) != 1:
            continue  # skip multi-range or sheet-level names

        sheet_title, coord = destinations[0]
        ws = wb[sheet_title] if sheet_title in wb.sheetnames else None
        if ws is None:
            continue

        # coord may be an absolute reference like "$B$5"
        coord_clean = coord.replace("$", "")
        try:
            cell = ws[coord_clean]
        except (KeyError, TypeError):
            continue

        if hasattr(cell, "value"):
            key = name.strip()
            data[key] = cell.value

    return data


def _read_assumptions_sheet(wb) -> dict:
    """
    Read from a sheet named 'Assumptions' (case-insensitive).

    Expects:
        Column A = assumption name (str)
        Column B = assumption value (numeric or str)

    Blank rows and rows where col A is not a string are skipped.
    """
    # Find the sheet case-insensitively
    target_ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == "assumptions":
            target_ws = wb[name]
            break

    if target_ws is None:
        return {}

    data: dict = {}
    for row in target_ws.iter_rows(min_row=1, values_only=True):
        if len(row) < 2:
            continue
        name_cell, value_cell = row[0], row[1]
        if not isinstance(name_cell, str):
            continue
        key = name_cell.strip()
        if not key:
            continue
        data[key] = value_cell

    return data


# ============================================================================
# VENDOR MODEL PARSER
# ============================================================================

# Row map for 1.1 Input_NTD: row_number → (key, type)
ROW_MAP_NTD: dict[int, tuple[str, type]] = {
    13: ("scenario_name", str), 15: ("hectare", float),
    31: ("res_mw_per_installation", float), 32: ("res_total_capacity_mw", float),
    36: ("res_lifetime_years", float),
    46: ("bess_active", bool), 47: ("bess_power_mw", float),
    48: ("bess_capacity_mwh", float), 52: ("bess_lifetime_years", float),
    66: ("production_p50_mwh", float), 67: ("production_p75_mwh", float),
    68: ("production_p90_mwh", float), 70: ("availability", float),
    71: ("grid_loss", float), 72: ("degradation", float),
    101: ("ppa_active", bool), 102: ("ppa_type", str),
    104: ("ppa_contracted_pct", float), 106: ("ppa_price", float),
    107: ("ppa_tenor", float), 109: ("ppa_escalation", float),
    120: ("goo_active", bool), 124: ("goo_price", float),
    139: ("tolling_active", bool), 140: ("bess_power_contracted", float),
    141: ("tolling_price", float), 143: ("tolling_duration", float),
    484: ("cit_rate", float), 485: ("vat_rate", float),
    711: ("shl_pct", float), 712: ("shl_margin", float),
}

# Curve blocks in 1.2 Input_M: (start_row, end_row, category, unit)
CURVE_BLOCKS = [
    (28, 66, "capture", "EUR/MWh"),
    (73, 109, "bess_revenue", "EUR/MW"),
    (111, 147, "pv_grid_revenue", "EUR"),
    (149, 185, "pv_ppa_revenue", "EUR"),
    (187, 223, "bess_opex", "EUR/MW"),
    (268, 270, "interest", "%"),
]


def parse_vendor_model(filepath: str) -> dict:
    """
    Parse full EE vendor model (.xlsx) into a structured dict.

    Reads tabs: 1.1 Input_NTD, 1.2 Input_M, 1.3 Input_A, 1.4 Global.
    Returns dict with keys: 'ntd', 'curves', 'annual', 'global_cfg'.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found: {filepath!r}")
    _, ext = os.path.splitext(filepath)
    if ext.lower() == ".xlsb":
        raise ValueError("Binary Excel (.xlsb) not supported. Save as .xlsx.")

    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    return {
        "ntd": _parse_input_ntd(wb),
        "curves": _parse_input_m(wb),
        "annual": _parse_input_a(wb),
        "global_cfg": _parse_global(wb),
    }


def _find_sheet(wb, *prefixes):
    """Find worksheet by prefix match (case-insensitive)."""
    for name in wb.sheetnames:
        for prefix in prefixes:
            if name.strip().lower().startswith(prefix.lower()):
                return wb[name]
    return None


def _parse_input_ntd(wb) -> dict:
    """Parse 1.1 Input_NTD: row-based, col F=value."""
    ws = _find_sheet(wb, "1.1")
    if ws is None:
        return {}
    data: dict = {}
    for row_num, (key, dtype) in ROW_MAP_NTD.items():
        val = ws.cell(row=row_num, column=6).value
        if val is None:
            data[key] = None
            continue
        try:
            if dtype == bool:
                data[key] = bool(val) if isinstance(val, (int, float)) else val == "Yes"
            elif dtype == float:
                data[key] = float(val) if val is not None else 0.0
            else:
                data[key] = str(val).strip()
        except (ValueError, TypeError):
            data[key] = val
    # Seasonality: rows 77-88
    data["seasonality"] = [
        float(ws.cell(row=r, column=6).value or 0.0) for r in range(77, 89)
    ]
    return data


def _parse_input_m(wb) -> dict:
    """Parse 1.2 Input_M: named monthly curves."""
    ws = _find_sheet(wb, "1.2")
    if ws is None:
        return {}
    curves: dict = {}
    for start_row, end_row, category, unit in CURVE_BLOCKS:
        for r in range(start_row, end_row + 1):
            name = ws.cell(row=r, column=4).value
            if not name or "[Placeholder" in str(name):
                continue
            name = str(name).strip()
            values: list[float] = []
            for c in range(6, min(getattr(ws, 'max_column', 600) + 1, 600)):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    break
                try:
                    values.append(float(v))
                except (ValueError, TypeError):
                    break
            if values:
                curves[name] = {
                    "name": name, "category": category, "unit": unit,
                    "frequency": "monthly", "values": values,
                }
    return curves


def _parse_input_a(wb) -> dict:
    """Parse 1.3 Input_A: annual curves."""
    ws = _find_sheet(wb, "1.3")
    if ws is None:
        return {}
    data: dict = {}
    for r in range(1, min(getattr(ws, 'max_row', 200) + 1, 200)):
        name = ws.cell(row=r, column=4).value
        if not name or not isinstance(name, str):
            continue
        values: list[float] = []
        for c in range(6, min(getattr(ws, 'max_column', 100) + 1, 100)):
            v = ws.cell(row=r, column=c).value
            if v is None:
                break
            try:
                values.append(float(v))
            except (ValueError, TypeError):
                break
        if values:
            data[name.strip()] = values
    return data


def _parse_global(wb) -> dict:
    """Parse 1.4 Global: config and lists."""
    ws = _find_sheet(wb, "1.4")
    if ws is None:
        return {}
    data: dict = {}
    for r in range(1, min(getattr(ws, 'max_row', 100) + 1, 100)):
        key = ws.cell(row=r, column=1).value
        val = ws.cell(row=r, column=2).value
        if key and isinstance(key, str):
            data[key.strip()] = val
    return data
