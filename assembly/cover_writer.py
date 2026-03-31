"""
assembly/cover_writer.py

Writes the Cover sheet with perspective-aware KPI blocks.

View selector in F5: 1 = Bank / Lender, 2 = IC (default), 3 = Audit.
KPI values reference Summary sheet via IF($F$5=...) formulas.
"""

from __future__ import annotations

from datetime import date
from typing import Any

from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from assembly.cell_mapper import COL_LABEL, COL_CONSTANT, COL_UNIT, COL_TOTAL
from assembly.excel_formatter import (
    FILL_COVER_HEADER, FILL_COVER_KPI, FILL_COVER_KEY_ROW,
    FILL_COL_HEADER, FILL_INPUT, FILL_SECTION_HEADER,
    FONT_WHITE, FONT_INPUT, FONT_IMPORT, FONT_EXPORT, FONT_DEFAULT,
)

# Cover column widths
COVER_COL_WIDTHS = {
    1: 1.3, 2: 1.3, 3: 1.3, 4: 1.3,  # A-D spacers
    5: 28.0,   # E labels
    6: 18.0,   # F values
    7: 22.0,   # G units/notes
    8: 2.0,    # H spacer
    9: 14.0,   # I FX label
    10: 10.0,  # J FX value
}

# Column indices
_E = COL_LABEL     # 5
_F = COL_CONSTANT  # 6
_G = COL_UNIT      # 7
_I = 9
_J = COL_TOTAL     # 10


# ============================================================================
# HELPERS
# ============================================================================

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color: str = FONT_DEFAULT, bold: bool = False,
          size: int = 10, italic: bool = False) -> Font:
    return Font(color=hex_color, bold=bold, name="Calibri", size=size,
                italic=italic)


def _find_summary_row(wb, label: str) -> str:
    """Find the row in Summary sheet where col E contains label.
    Returns a cell reference string like 'Summary!$F$12', or '""' if not found.
    """
    if "Summary" not in wb.sheetnames:
        return '""'
    ws = wb["Summary"]
    for row in ws.iter_rows(min_col=_E, max_col=_E):
        for cell in row:
            if cell.value and label.lower() in str(cell.value).lower():
                return f"Summary!$F${cell.row}"
    return '""'


def _get_capacity_mw(config) -> float | None:
    """Try to extract capacity from config."""
    if getattr(config, 'opex_pv', None) is not None:
        return getattr(config.opex_pv, 'capacity_mwp', None)
    if getattr(config, 'opex_wind', None) is not None:
        return getattr(config.opex_wind, 'capacity_mwp', None)
    if getattr(config, 'rev_bess', None) is not None:
        return getattr(config.rev_bess, 'power_MW', None)
    return None


def _get_cod_date(config, result) -> date | None:
    """Compute COD from capex construction parameters."""
    capex = getattr(config, 'capex', None)
    if capex is None:
        return None
    cp = getattr(capex, 'construction_periods', 0)
    cs = getattr(capex, 'construction_start_period', 0)
    cod_period = cs + cp
    sy, sm = result.start_year, result.start_month
    total_months = (sm - 1) + cod_period
    year = sy + total_months // 12
    month = total_months % 12 + 1
    return date(year, month, 1)


# ============================================================================
# BANNER (rows 1-3)
# ============================================================================

def _write_banner(ws, result, config) -> None:
    """Teal banner with project name, tech/market, version, FX rate."""
    for row in range(1, 4):
        for col in range(1, 11):  # A through J
            cell = ws.cell(row=row, column=col)
            cell.fill = _fill(FILL_COVER_HEADER)
            cell.font = _font(FONT_WHITE, bold=True, size=12)

    # E1: project name — F1F9 uses Arial 20pt bold
    ws.cell(row=1, column=_E, value=result.project_name)
    ws.cell(row=1, column=_E).font = Font(name="Arial", size=20, bold=True, color=FONT_WHITE)
    tech = getattr(config, 'technology', 'PV')
    market = getattr(config, 'market', 'DK')
    ws.cell(row=2, column=_E, value=f"{tech} | {market}")
    ws.cell(row=3, column=_E, value="EE Model Library v1.0")

    # FX rate input cell
    ws.cell(row=2, column=_I, value="DKK/EUR rate:").font = _font(
        FONT_WHITE, bold=False, size=10)
    fx_cell = ws.cell(row=2, column=_J, value=7.46)
    fx_cell.fill = _fill(FILL_INPUT)
    fx_cell.font = _font(FONT_INPUT, bold=True)
    fx_cell.number_format = '0.00'


# ============================================================================
# VIEW SELECTOR (row 5)
# ============================================================================

def _write_view_selector(ws, view_default: int) -> None:
    ws.cell(row=5, column=_E, value="Active view:").font = _font(bold=True)
    sel = ws.cell(row=5, column=_F, value=view_default)
    sel.fill = _fill(FILL_INPUT)
    sel.font = _font(FONT_INPUT, bold=True)
    sel.comment = Comment("1 = Bank / Lender    2 = IC    3 = Audit",
                          "EE Model Library")
    ws.cell(row=5, column=_G,
            value="(1=Bank  2=IC  3=Audit)").font = _font(
        "A6A6A6", italic=True)


# ============================================================================
# PROJECT DETAILS (rows 7-14)
# ============================================================================

def _write_project_details(ws, result, config) -> None:
    # Row 7: header
    for col in range(_E, _J + 1):
        ws.cell(row=7, column=col).fill = _fill(FILL_COL_HEADER)
        ws.cell(row=7, column=col).font = _font(FONT_WHITE, bold=True)
    ws.cell(row=7, column=_E, value="Project details")

    details = [
        (8,  "Project name",   result.project_name, ""),
        (9,  "Technology",     getattr(config, 'technology', 'PV'), ""),
        (10, "Market",         getattr(config, 'market', 'DK'), ""),
    ]

    cap = _get_capacity_mw(config)
    details.append((11, "Capacity", cap if cap else "N/A", "MW AC"))

    cod = _get_cod_date(config, result)
    details.append((12, "COD", cod if cod else "N/A", ""))

    life_years = round(result.periods / 12, 1)
    details.append((13, "Project life", life_years, "years"))
    details.append((14, "Model version", "1.0", ""))

    for row, label, value, unit in details:
        ws.cell(row=row, column=_E, value=label).font = _font(bold=False)
        cell = ws.cell(row=row, column=_F, value=value)
        if isinstance(value, date):
            cell.number_format = "DD-MMM-YY"
        ws.cell(row=row, column=_G, value=unit)


# ============================================================================
# KPI BLOCK (rows 16-27)
# ============================================================================

# Each KPI row: (bank_label, bank_unit, ic_label, ic_unit, audit_label, audit_unit)
_KPI_DEFS = [
    # row 17
    ("Min DSCR",       "x",    "Project IRR",        "%",    "Model checks",  "#"),
    # row 18
    ("Avg DSCR",       "x",    "Equity IRR",         "%",    "BS balance",    "OK/FAIL"),
    # row 19
    ("LLCR",           "x",    "IRR spread vs WACC", "bps",  "DSCR breaches", "#"),
    # row 20
    ("Debt quantum",   "DKKm", "Project NPV",        "DKKm", "Sum check",     "OK/FAIL"),
    # row 21
    ("Gearing",        "%",    "Project NPV",        "EURm", "",              ""),
    # row 22
    ("Debt tenor",     "years","Project NPV",        "EUR/kW","",             ""),
    # row 23
    ("",               "",     "Total revenue",      "DKKm", "",              ""),
    # row 24
    ("",               "",     "Total OPEX",         "DKKm", "",              ""),
    # row 25
    ("",               "",     "Total CAPEX",        "DKKm", "",              ""),
    # row 26
    ("",               "",     "EBITDA margin",      "%",    "",              ""),
    # row 27
    ("",               "",     "WACC",               "%",    "",              ""),
]

# Summary label mappings for each KPI value
# (bank_summary_label, ic_summary_label, audit_summary_label)
_KPI_SUMMARY_LABELS = [
    ("Min DSCR",       "Project IRR",   None),        # row 17
    (None,             "Equity IRR",    None),        # row 18
    (None,             None,            None),        # row 19 — computed
    (None,             "Project NPV",   None),        # row 20
    (None,             "Project NPV",   None),        # row 21 — EUR conversion
    (None,             "Project NPV",   None),        # row 22 — EUR/kW
    (None,             None,            None),        # row 23
    (None,             None,            None),        # row 24
    ("Total CAPEX",    "Total CAPEX",   None),        # row 25
    (None,             None,            None),        # row 26
    ("WACC",           "WACC",          None),        # row 27
]


def _write_kpi_block(ws, result, config, wb) -> None:
    """Write KPI header (row 16) and KPI rows (17-27) with IF formulas."""
    # Row 16: KPI block header
    for col in range(_E, _J + 1):
        ws.cell(row=16, column=col).fill = _fill(FILL_COL_HEADER)
        ws.cell(row=16, column=col).font = _font(FONT_WHITE, bold=True)
    ws.cell(row=16, column=_E,
            value='=IF($F$5=1,"Bank / Lender KPIs",IF($F$5=2,"IC KPIs","Model Audit"))')

    # Resolve summary references
    irr_ref = _find_summary_row(wb, "Project IRR")
    eirr_ref = _find_summary_row(wb, "Equity IRR")
    wacc_ref = _find_summary_row(wb, "WACC")
    npv_ref = _find_summary_row(wb, "Project NPV")
    dscr_ref = _find_summary_row(wb, "Min DSCR")
    capex_ref = _find_summary_row(wb, "Total CAPEX")

    cap = _get_capacity_mw(config)
    fx_ref = "Cover!$J$2"

    for i, (bank_l, bank_u, ic_l, ic_u, audit_l, audit_u) in enumerate(_KPI_DEFS):
        row = 17 + i

        # Background fill
        for col in range(_E, _G + 1):
            ws.cell(row=row, column=col).fill = _fill(FILL_COVER_KPI)

        # Label formula (col E)
        bl = bank_l if bank_l else ""
        il = ic_l if ic_l else ""
        al = audit_l if audit_l else ""
        ws.cell(row=row, column=_E,
                value=f'=IF($F$5=1,"{bl}",IF($F$5=2,"{il}","{al}"))')

        # Unit formula (col G)
        bu = bank_u if bank_u else ""
        iu = ic_u if ic_u else ""
        au = audit_u if audit_u else ""
        ws.cell(row=row, column=_G,
                value=f'=IF($F$5=1,"{bu}",IF($F$5=2,"{iu}","{au}"))')

        # Value formula (col F) — each row is specific
        bank_val = '""'
        ic_val = '""'
        audit_val = '""'

        if i == 0:  # Min DSCR / Project IRR / Model checks
            bank_val = dscr_ref
            ic_val = irr_ref
        elif i == 1:  # Avg DSCR / Equity IRR / BS balance
            ic_val = eirr_ref
        elif i == 2:  # LLCR / IRR spread vs WACC / DSCR breaches
            if irr_ref != '""' and wacc_ref != '""':
                ic_val = f"({irr_ref}-{wacc_ref})*10000"
        elif i == 3:  # Debt quantum / Project NPV DKKm / Sum check
            if npv_ref != '""':
                ic_val = f"{npv_ref}/1000"
        elif i == 4:  # Gearing / Project NPV EURm
            if npv_ref != '""':
                ic_val = f"{npv_ref}/1000/{fx_ref}"
        elif i == 5:  # Debt tenor / Project NPV EUR/kW
            if npv_ref != '""' and cap:
                ic_val = f"{npv_ref}/{fx_ref}/({cap}*1000)*1000000"
        elif i == 8:  # Total CAPEX
            bank_val = capex_ref if capex_ref != '""' else '""'
            if capex_ref != '""':
                ic_val = f"{capex_ref}/1000"
        elif i == 10:  # WACC
            bank_val = wacc_ref
            ic_val = wacc_ref

        ws.cell(row=row, column=_F,
                value=f'=IF($F$5=1,{bank_val},IF($F$5=2,{ic_val},{audit_val}))')


# ============================================================================
# COLOR CODING KEY (rows 29-34)
# ============================================================================

def _write_color_key(ws) -> None:
    """Write color coding legend rows 29-34."""
    ws.cell(row=29, column=_E, value="Color coding key").font = _font(bold=True)

    key_rows = [
        # (row, swatch_fill, swatch_font, label, description)
        (30, FILL_INPUT,          None,         "Input cell",      "Editable assumption"),
        (31, None,                FONT_IMPORT,  "Imported value",  "Pulled from calc sheet"),
        (32, None,                FONT_EXPORT,  "Exported value",  "Used by other sheets"),
        (33, FILL_SECTION_HEADER, FONT_WHITE,   "Section header",  "Group label"),
        (34, None,                None,         "Subtotal",        "Thin top+bottom border"),
    ]

    for row, swatch_fill, swatch_font_color, label, desc in key_rows:
        for col in range(_E, _G + 1):
            ws.cell(row=row, column=col).fill = _fill(FILL_COVER_KEY_ROW)

        swatch = ws.cell(row=row, column=_E)
        if swatch_fill:
            swatch.fill = _fill(swatch_fill)
            swatch.value = "  ■"
            if swatch_fill == FILL_SECTION_HEADER:
                swatch.font = _font(FONT_WHITE, bold=True)
        elif swatch_font_color:
            swatch.value = "  ■"
            swatch.font = _font(swatch_font_color, bold=True)
        else:
            # Border swatch
            swatch.value = "  —"
            swatch.border = Border(top=Side(border_style="thin"),
                                   bottom=Side(border_style="thin"))

        ws.cell(row=row, column=_F, value=label)
        ws.cell(row=row, column=_G, value=desc)


# ============================================================================
# BRANDING (row 36)
# ============================================================================

def _write_branding(ws) -> None:
    ws.cell(row=36, column=_E,
            value="European Energy A/S  |  EE Model Library").font = _font(
        "A6A6A6", size=8, italic=True)


# ============================================================================
# PUBLIC API
# ============================================================================

def write_cover(ws, result, config, wb,
                view_selector_default: int = 2) -> None:
    """Write the full Cover sheet.
    wb is passed so _find_summary_row() can read the Summary sheet.
    """
    _write_banner(ws, result, config)
    _write_view_selector(ws, view_selector_default)
    _write_project_details(ws, result, config)
    _write_kpi_block(ws, result, config, wb)
    _write_color_key(ws)
    _write_branding(ws)

    # Column widths
    for col_idx, width in COVER_COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
