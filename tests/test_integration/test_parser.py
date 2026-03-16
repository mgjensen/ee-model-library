"""Tests for assembly/parser.py — .xlsx assumption file reader."""

import os
import pytest
import openpyxl

from assembly.parser import parse_xlsx, _read_named_ranges, _read_assumptions_sheet


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_assumptions_xlsx(tmp_path, rows: list[tuple]) -> str:
    """Create a minimal xlsx with an 'Assumptions' sheet populated from rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    for row in rows:
        ws.append(row)
    path = str(tmp_path / "assumptions.xlsx")
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Assumptions sheet — round-trip
# ---------------------------------------------------------------------------

def test_parse_assumptions_sheet_basic(tmp_path):
    path = _make_assumptions_xlsx(tmp_path, [
        ("wacc", 0.07),
        ("inflation_rate", 0.025),
        ("facility_DKKk", 450_000.0),
    ])
    data = parse_xlsx(path)
    assert data["wacc"] == pytest.approx(0.07)
    assert data["inflation_rate"] == pytest.approx(0.025)
    assert data["facility_DKKk"] == pytest.approx(450_000.0)


def test_parse_assumptions_sheet_string_value(tmp_path):
    path = _make_assumptions_xlsx(tmp_path, [
        ("market", "DK"),
        ("technology", "PV"),
    ])
    data = parse_xlsx(path)
    assert data["market"] == "DK"
    assert data["technology"] == "PV"


def test_parse_assumptions_sheet_skips_blank_rows(tmp_path):
    path = _make_assumptions_xlsx(tmp_path, [
        ("key1", 1.0),
        (None, None),       # blank row
        ("", 99.0),          # empty key
        ("key2", 2.0),
    ])
    data = parse_xlsx(path)
    assert "key1" in data
    assert "key2" in data
    assert len(data) == 2


def test_parse_assumptions_sheet_skips_numeric_keys(tmp_path):
    """Rows where col A is a number (not a string) should be skipped."""
    path = _make_assumptions_xlsx(tmp_path, [
        (1, 100.0),         # numeric key — should be skipped
        ("valid", 200.0),
    ])
    data = parse_xlsx(path)
    assert 1 not in data
    assert "valid" in data


def test_parse_assumptions_sheet_strips_whitespace(tmp_path):
    path = _make_assumptions_xlsx(tmp_path, [
        ("  wacc  ", 0.07),
    ])
    data = parse_xlsx(path)
    assert "wacc" in data
    assert "  wacc  " not in data


def test_parse_assumptions_case_insensitive_sheet_name(tmp_path):
    """Sheet named 'ASSUMPTIONS' should still be read."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ASSUMPTIONS"
    ws.append(("rate", 0.05))
    path = str(tmp_path / "upper.xlsx")
    wb.save(path)
    data = parse_xlsx(path)
    assert "rate" in data


def test_parse_no_assumptions_sheet_returns_empty(tmp_path):
    """Workbook without 'Assumptions' sheet returns empty dict."""
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb["Sheet1"].append(("value", 42))
    path = str(tmp_path / "no_assumptions.xlsx")
    wb.save(path)
    data = parse_xlsx(path)
    assert data == {}


# ---------------------------------------------------------------------------
# Named ranges
# ---------------------------------------------------------------------------

def test_parse_named_ranges(tmp_path):
    """Named ranges pointing to single cells should be extracted."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    ws["B5"] = 0.07

    # Define a named range
    from openpyxl.workbook.defined_name import DefinedName
    defn = DefinedName(name="wacc_named", attr_text="Assumptions!$B$5")
    wb.defined_names.add(defn)

    path = str(tmp_path / "named.xlsx")
    wb.save(path)

    data = parse_xlsx(path)
    assert "wacc_named" in data
    assert data["wacc_named"] == pytest.approx(0.07)


def test_named_ranges_and_assumptions_merged(tmp_path):
    """Both strategies run; Assumptions sheet values override named ranges on conflict."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    ws["A1"] = "rate"
    ws["B1"] = 0.10        # Assumptions sheet: rate = 0.10
    ws["B5"] = 0.07        # named range: rate = 0.07 (overridden by Assumptions)

    from openpyxl.workbook.defined_name import DefinedName
    defn = DefinedName(name="rate", attr_text="Assumptions!$B$5")
    wb.defined_names.add(defn)

    path = str(tmp_path / "merged.xlsx")
    wb.save(path)

    data = parse_xlsx(path)
    # Assumptions sheet wins (applied second in parse_xlsx)
    assert data["rate"] == pytest.approx(0.10)


# ---------------------------------------------------------------------------
# Error cases
# ---------------------------------------------------------------------------

def test_missing_file_raises_file_not_found():
    with pytest.raises(FileNotFoundError, match="not found"):
        parse_xlsx("/nonexistent/path/file.xlsx")


def test_xlsb_raises_value_error(tmp_path):
    """Passing a .xlsb path should raise ValueError with instructions."""
    fake_xlsb = str(tmp_path / "model.xlsb")
    # Create a dummy file so the extension check is reached
    with open(fake_xlsb, "w") as f:
        f.write("dummy")
    with pytest.raises(ValueError, match=r"(?i)xlsb|xlsx"):
        parse_xlsx(fake_xlsb)


def test_xlsb_missing_file_raises_file_not_found():
    """A missing .xlsb should raise FileNotFoundError, not ValueError."""
    with pytest.raises(FileNotFoundError):
        parse_xlsx("/nonexistent/model.xlsb")


# ---------------------------------------------------------------------------
# _read_assumptions_sheet directly
# ---------------------------------------------------------------------------

def test_read_assumptions_sheet_returns_dict(tmp_path):
    path = _make_assumptions_xlsx(tmp_path, [("x", 1.0), ("y", 2.0)])
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    result = _read_assumptions_sheet(wb)
    assert result == {"x": 1.0, "y": 2.0}


def test_read_assumptions_sheet_no_sheet_returns_empty(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "NotAssumptions"
    result = _read_assumptions_sheet(wb)
    assert result == {}
