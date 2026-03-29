"""Integration tests for assembly/excel_writer.py — openpyxl workbook writer."""

import os
import math
import pytest
import openpyxl

from assembly.engine import (
    ProjectConfig,
    TimelineConfig,
    TaxConfig,
    StatementConfig,
    run,
)
from assembly.excel_writer import write_workbook, SHEETS, _period_to_label
from assembly.cell_mapper import (
    get_row, period_col, col_letter,
    COL_LABEL, COL_UNIT, COL_CONSTANT, COL_TOTAL, COL_PERIOD_0,
)
import modules.revenue.REV_001 as REV_001
import modules.costs.OPEX_001 as OPEX_001
import modules.capex.CAPEX_001 as CAPEX_001
import modules.debt.DEBT_001 as DEBT_001


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _cr(annual=100.0):
    return OPEX_001.ComponentRate(annual_DKKk=annual)


def _zero_cr():
    return _cr(0.0)


@pytest.fixture
def minimal_result(tmp_path):
    n = 24
    config = ProjectConfig(
        project_name="Writer Test",
        timeline=TimelineConfig(periods=n, start_year=2026, start_month=1),
        rev_pv=REV_001.Inputs(
            periods=n, start_year=2026, start_month=1,
            net_production_MWh=[500.0] * n,
            wholesale_price_DKK_per_MWh=[400.0] * n,
            capture_price_DKK_per_MWh=[380.0] * n,
            goo_price_DKK_per_MWh=[5.0] * n,
        ),
        opex_pv=OPEX_001.Inputs(
            periods=n, start_year=2026, start_month=1,
            capacity_mwp=10.0,
            om=_cr(500.0), commercial_management=_zero_cr(),
            inverter_replacement=_zero_cr(), insurance=_cr(100.0),
            other=_zero_cr(),
        ),
        capex=CAPEX_001.Inputs(
            periods=n,
            construction_start_period=0, construction_periods=12,
            drawdown_profile=[1.0 / 12] * 12,
            epc_DKKk=40_000.0, grid_connection_DKKk=5_000.0,
            development_DKKk=2_000.0, land_DKKk=0.0,
            contingency_pct=0.05, other_DKKk=0.0,
        ),
        debt=DEBT_001.Inputs(
            periods=n, start_year=2026, start_month=1,
            facility=35_000.0, all_in_rate=0.05,
            repayment_type="annuity", tenor_months=12,
            drawdowns=[35_000.0 / 12] * 12 + [0.0] * 12,
        ),
        tax=TaxConfig(country="DK"),
        statements=StatementConfig(
            opening_cash_DKKk=1_000.0,
            opening_contributed_equity_DKKk=15_000.0,
        ),
    )
    result = run(config)
    return result, config


@pytest.fixture
def written_wb(minimal_result, tmp_path):
    result, config = minimal_result
    out_path = str(tmp_path / "test_out.xlsx")
    write_workbook(result, config, out_path)
    return openpyxl.load_workbook(out_path), result, config, out_path


# ---------------------------------------------------------------------------
# File creation
# ---------------------------------------------------------------------------

def test_creates_file(minimal_result, tmp_path):
    result, config = minimal_result
    out_path = str(tmp_path / "output.xlsx")
    write_workbook(result, config, out_path)
    assert os.path.exists(out_path)


def test_file_is_valid_xlsx(written_wb):
    wb, *_ = written_wb
    assert wb is not None


# ---------------------------------------------------------------------------
# Sheets present
# ---------------------------------------------------------------------------

def test_all_sheets_present(written_wb):
    wb, *_ = written_wb
    expected = {"Assumptions", "Inputs"} | set(SHEETS)
    assert set(wb.sheetnames) == expected


def test_sheet_order(written_wb):
    wb, *_ = written_wb
    expected = ["Assumptions", "Inputs"] + SHEETS
    assert wb.sheetnames == expected


def test_no_old_statements_sheet(written_wb):
    wb, *_ = written_wb
    assert "Statements" not in wb.sheetnames


# ---------------------------------------------------------------------------
# Header rows — new column layout
# ---------------------------------------------------------------------------

def test_header_row1_contains_sheet_name(written_wb):
    wb, result, *_ = written_wb
    ws = wb["Revenue"]
    assert ws.cell(row=1, column=COL_LABEL).value == "Revenue"


def test_header_row5_description(written_wb):
    wb, *_ = written_wb
    ws = wb["Revenue"]
    assert ws.cell(row=5, column=COL_LABEL).value == "Description"


def test_header_row5_constant(written_wb):
    wb, *_ = written_wb
    ws = wb["Revenue"]
    assert ws.cell(row=5, column=COL_CONSTANT).value == "Constant"


def test_header_row5_unit(written_wb):
    wb, *_ = written_wb
    ws = wb["Revenue"]
    assert ws.cell(row=5, column=COL_UNIT).value == "Unit"


def test_header_row5_total(written_wb):
    wb, *_ = written_wb
    ws = wb["Revenue"]
    assert ws.cell(row=5, column=COL_TOTAL).value == "Total / avg."


def test_freeze_panes_l7(written_wb):
    wb, *_ = written_wb
    for sheet_name in ["Revenue", "Costs", "Debt", "FS_Monthly"]:
        ws = wb[sheet_name]
        assert str(ws.freeze_panes) == "L7", f"{sheet_name} freeze panes wrong"


def test_hidden_columns(written_wb):
    wb, *_ = written_wb
    ws = wb["Revenue"]
    assert ws.column_dimensions["H"].hidden is True
    assert ws.column_dimensions["I"].hidden is True


# ---------------------------------------------------------------------------
# Data rows — Revenue sheet (new column positions)
# ---------------------------------------------------------------------------

def test_rev001_net_revenue_written(written_wb):
    wb, result, *_ = written_wb
    ws = wb["Revenue"]
    row = get_row("REV_001", "net_revenue")
    rev_out = result.outputs["REV_001"]
    expected = rev_out.net_revenue[0]
    actual = ws.cell(row=row, column=period_col(0)).value
    assert actual == pytest.approx(expected, rel=1e-6)


def test_rev001_label_in_col_e(written_wb):
    wb, *_ = written_wb
    ws = wb["Revenue"]
    row = get_row("REV_001", "net_revenue")
    label = ws.cell(row=row, column=COL_LABEL).value
    assert label is not None and "revenue" in label.lower()


def test_rev001_units_in_col_g(written_wb):
    wb, *_ = written_wb
    ws = wb["Revenue"]
    row = get_row("REV_001", "net_revenue")
    units = ws.cell(row=row, column=COL_UNIT).value
    assert units == "kEUR"  # Default currency label


def test_rev001_total_in_col_j(written_wb):
    wb, result, *_ = written_wb
    ws = wb["Revenue"]
    row = get_row("REV_001", "net_revenue")
    total = ws.cell(row=row, column=COL_TOTAL).value
    assert total is not None and isinstance(total, (int, float))


# ---------------------------------------------------------------------------
# Data rows — Costs sheet
# ---------------------------------------------------------------------------

def test_opex_total_written(written_wb):
    wb, result, *_ = written_wb
    ws = wb["Costs"]
    row = get_row("OPEX_001", "total_opex")
    opex_out = result.outputs["OPEX_001"]
    expected = opex_out.total_opex[0]
    actual = ws.cell(row=row, column=period_col(0)).value
    assert actual == pytest.approx(expected, rel=1e-6)


def test_capex_cumulative_written(written_wb):
    wb, result, *_ = written_wb
    ws = wb["Costs"]
    row = get_row("CAPEX_001", "cumulative_capex")
    capex_out = result.outputs["CAPEX_001"]
    for p in range(3):
        expected = capex_out.cumulative_capex[p]
        actual = ws.cell(row=row, column=period_col(p)).value
        assert actual == pytest.approx(expected, rel=1e-6)


# ---------------------------------------------------------------------------
# Data rows — FS_Monthly sheet (replaces Statements)
# ---------------------------------------------------------------------------

def test_pl_net_income_written(written_wb):
    wb, result, *_ = written_wb
    ws = wb["FS_Monthly"]
    row = get_row("PL_001", "net_income")
    actual = ws.cell(row=row, column=period_col(0)).value
    # May be a formula string (Layer 1) or a float (Layer 2)
    assert actual is not None
    if isinstance(actual, str):
        assert actual.startswith("="), f"Expected formula, got: {actual!r}"
    else:
        pl_out = result.outputs["PL_001"]
        expected = pl_out.net_income[0]
        assert actual == pytest.approx(expected, rel=1e-6)


def test_irr_cumulative_pfcf_written(written_wb):
    wb, result, *_ = written_wb
    if "IRR_001" not in result.outputs:
        pytest.skip("IRR_001 not in outputs")
    ws = wb["FS_Monthly"]
    row = get_row("IRR_001", "cumulative_pfcf")
    irr_out = result.outputs["IRR_001"]
    expected = irr_out.cumulative_pfcf[0]
    actual = ws.cell(row=row, column=period_col(0)).value
    assert actual == pytest.approx(expected, rel=1e-6)


# ---------------------------------------------------------------------------
# FS_Annual sheet
# ---------------------------------------------------------------------------

def test_fs_annual_exists(written_wb):
    wb, *_ = written_wb
    assert "FS_Annual" in wb.sheetnames


def test_fs_annual_has_year_in_row2(written_wb):
    wb, *_ = written_wb
    ws = wb["FS_Annual"]
    yr = ws.cell(row=2, column=COL_PERIOD_0).value
    assert isinstance(yr, int) and yr == 2026


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def test_summary_sheet_has_project_name(written_wb):
    wb, result, *_ = written_wb
    ws = wb["Summary"]
    assert ws.cell(row=1, column=COL_LABEL).value == result.project_name


def test_summary_has_irr_entry(written_wb):
    wb, result, *_ = written_wb
    if "IRR_001" not in result.outputs:
        pytest.skip("IRR_001 not in outputs")
    ws = wb["Summary"]
    labels = [ws.cell(row=r, column=COL_LABEL).value for r in range(1, 30)]
    assert any(l and "irr" in str(l).lower() for l in labels)


def test_summary_has_npv_entry(written_wb):
    wb, result, *_ = written_wb
    if "IRR_001" not in result.outputs:
        pytest.skip("IRR_001 not in outputs")
    ws = wb["Summary"]
    labels = [ws.cell(row=r, column=COL_LABEL).value for r in range(1, 30)]
    assert any(l and "npv" in str(l).lower() for l in labels)


# ---------------------------------------------------------------------------
# Disabled module rows are empty
# ---------------------------------------------------------------------------

def test_disabled_module_rows_have_no_data(tmp_path):
    """REV_002 disabled → its rows in Revenue sheet have no numeric values."""
    n = 12
    config = ProjectConfig(
        project_name="PV Only",
        timeline=TimelineConfig(periods=n, start_year=2026, start_month=1),
        rev_pv=REV_001.Inputs(
            periods=n, start_year=2026, start_month=1,
            net_production_MWh=[500.0] * n,
            wholesale_price_DKK_per_MWh=[400.0] * n,
            capture_price_DKK_per_MWh=[380.0] * n,
            goo_price_DKK_per_MWh=[5.0] * n,
        ),
        # rev_bess = None → disabled
    )
    result = run(config)
    out_path = str(tmp_path / "pv_only.xlsx")
    write_workbook(result, config, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["Revenue"]
    row = get_row("REV_002", "net_revenue")
    # Period columns should be None (no data written)
    for p in range(n):
        val = ws.cell(row=row, column=period_col(p)).value
        assert val is None, f"Expected None at REV_002 row period {p}, got {val}"


# ---------------------------------------------------------------------------
# Cover sheet
# ---------------------------------------------------------------------------

def test_cover_sheet_has_project_name(written_wb):
    wb, result, *_ = written_wb
    ws = wb["Cover"]
    assert ws.cell(row=1, column=COL_LABEL).value == result.project_name


# ---------------------------------------------------------------------------
# _period_to_label
# ---------------------------------------------------------------------------

def test_period_to_label_jan_start():
    assert _period_to_label(2026, 1, 0) == "Jan-2026"


def test_period_to_label_rolls_over_year():
    assert _period_to_label(2026, 1, 12) == "Jan-2027"


def test_period_to_label_mid_year():
    assert _period_to_label(2026, 6, 6) == "Dec-2026"


def test_period_to_label_dec_rolls_to_jan():
    assert _period_to_label(2026, 12, 1) == "Jan-2027"
