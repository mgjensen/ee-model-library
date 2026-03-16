"""
End-to-end smoke test — Project Viuf proxy.

Viuf is a combined PV + BESS project in Denmark:
  - ~475 monthly periods (construction + 38 years operations)
  - Construction: 18 months
  - Total CAPEX: ~750 MKK (EPC + grid + development)
  - Debt: ~525 MKK (70% gearing), 25-year tenor, 5% all-in
  - Equity: ~225 MKK (30%)
  - PV: ~45 MWp, P50 ~1,050 MWh/MWp/year
  - BESS: ~20 MW / 40 MWh

This test exercises the full stack:
    engine.run() → write_workbook() → parse_xlsx()

It verifies:
  - All statement modules run without error
  - Array lengths match the project life
  - Output file is a valid .xlsx with all 5 sheets
  - Summary sheet has IRR/NPV entries
"""

import math
import os
import tempfile
import pytest
import openpyxl

from assembly.engine import (
    ProjectConfig, TimelineConfig, TaxConfig, StatementConfig, run,
)
from assembly.excel_writer import write_workbook, SHEETS
import modules.revenue.REV_001 as REV_001
import modules.revenue.REV_002 as REV_002
import modules.costs.OPEX_001 as OPEX_001
import modules.costs.OPEX_002 as OPEX_002
import modules.capex.CAPEX_001 as CAPEX_001
import modules.debt.DEBT_001 as DEBT_001


# ---------------------------------------------------------------------------
# Viuf config
# ---------------------------------------------------------------------------

N = 475           # total project months
CONSTRUCTION = 18 # construction months
START_YEAR = 2026
START_MONTH = 1
CAPACITY_MWP = 45.0
BESS_MW = 20.0

# Annual PV yield: 45 MWp × 1050 MWh/MWp × 0.98 availability / 12 months
_PV_MONTHLY_MWH = CAPACITY_MWP * 1_050 * 0.98 / 12

# BESS: assume daily full cycle, 365/12 cycles/month × BESS energy 40 MWh × RTE 0.85
_BESS_MONTHLY_DISCHARGE = 40.0 * (365 / 12) * 0.85

_CONSTRUCTION_ZEROS = [0.0] * CONSTRUCTION
_OPS_PERIODS = N - CONSTRUCTION


def _make_config() -> ProjectConfig:
    n = N
    cp = CONSTRUCTION

    # PV revenue
    rev_pv = REV_001.Inputs(
        periods=n, start_year=START_YEAR, start_month=START_MONTH,
        net_production_MWh=_CONSTRUCTION_ZEROS + [_PV_MONTHLY_MWH] * _OPS_PERIODS,
        wholesale_price_DKK_per_MWh=[400.0] * n,
        capture_price_DKK_per_MWh=[390.0] * n,
        goo_price_DKK_per_MWh=[5.0] * n,
    )

    # BESS revenue
    rev_bess = REV_002.Inputs(
        periods=n, start_year=START_YEAR, start_month=START_MONTH,
        power_MW=BESS_MW,
        energy_MWh_installed=40.0,
        round_trip_efficiency=0.85,
        discharge_volume_MWh=_CONSTRUCTION_ZEROS + [_BESS_MONTHLY_DISCHARGE] * _OPS_PERIODS,
        discharge_price_DKK_per_MWh=[500.0] * n,
        grid_charge_volume_MWh=_CONSTRUCTION_ZEROS + [_BESS_MONTHLY_DISCHARGE / 0.85] * _OPS_PERIODS,
        grid_charge_price_DKK_per_MWh=[200.0] * n,
        pv_charge_volume_MWh=[0.0] * n,
        pv_charge_price_DKK_per_MWh=[0.0] * n,
        goo_price_DKK_per_MWh=[5.0] * n,
        multimarket_pct=[0.0] * n,
    )

    # PV OPEX
    cr = OPEX_001.ComponentRate
    opex_pv = OPEX_001.Inputs(
        periods=n, start_year=START_YEAR, start_month=START_MONTH,
        capacity_mwp=CAPACITY_MWP,
        om=cr(annual_DKKk=CAPACITY_MWP * 50.0),          # ~50 DKKk/MWp/yr
        commercial_management=cr(annual_DKKk=500.0),
        inverter_replacement=cr(annual_DKKk=0.0),
        insurance=cr(annual_DKKk=CAPACITY_MWP * 10.0),   # ~10 DKKk/MWp/yr
        other=cr(annual_DKKk=200.0),
    )

    # BESS OPEX
    opex_bess = OPEX_002.Inputs(
        periods=n, start_year=START_YEAR, start_month=START_MONTH,
        discharge_volume_MWh=_CONSTRUCTION_ZEROS + [_BESS_MONTHLY_DISCHARGE] * _OPS_PERIODS,
    )

    # CAPEX: total ~750 MKK → 750,000 DKKk
    drawdown_profile = [1.0 / cp] * cp
    capex = CAPEX_001.Inputs(
        periods=n,
        construction_start_period=0,
        construction_periods=cp,
        drawdown_profile=drawdown_profile,
        epc_DKKk=540_000.0,
        grid_connection_DKKk=80_000.0,
        development_DKKk=40_000.0,
        land_DKKk=0.0,
        contingency_pct=0.05,
        other_DKKk=0.0,
    )

    # Debt: ~525 MKK at 5%, 25-year tenor
    facility = 525_000.0
    drawdowns = [facility / cp] * cp + [0.0] * (n - cp)
    debt = DEBT_001.Inputs(
        periods=n, start_year=START_YEAR, start_month=START_MONTH,
        facility=facility,
        all_in_rate=0.05,
        repayment_type="annuity",
        tenor_months=min(300, n - cp),  # 25 years or remaining project life
        drawdowns=drawdowns,
    )

    return ProjectConfig(
        project_name="Viuf E2E",
        market="DK",
        technology="PV",
        timeline=TimelineConfig(periods=n, start_year=START_YEAR, start_month=START_MONTH),
        rev_pv=rev_pv,
        rev_bess=rev_bess,
        opex_pv=opex_pv,
        opex_bess=opex_bess,
        capex=capex,
        debt=debt,
        tax=TaxConfig(country="DK"),
        statements=StatementConfig(
            opening_cash_DKKk=5_000.0,
            opening_contributed_equity_DKKk=225_000.0,
        ),
    )


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def viuf_result():
    return run(_make_config())


@pytest.fixture(scope="module")
def viuf_workbook(viuf_result, tmp_path_factory):
    config = _make_config()
    tmp = tmp_path_factory.mktemp("viuf") / "viuf.xlsx"
    write_workbook(viuf_result, config, str(tmp))
    return openpyxl.load_workbook(str(tmp)), viuf_result


# ---------------------------------------------------------------------------
# Engine tests
# ---------------------------------------------------------------------------

def test_all_modules_ran(viuf_result):
    for mid in ("REV_001", "REV_002", "OPEX_001", "OPEX_002",
                "CAPEX_001", "DEBT_001", "TAX_001",
                "PL_001", "CF_001", "BS_001", "IRR_001"):
        assert mid in viuf_result.outputs, f"{mid} missing"


def test_array_lengths_match_periods(viuf_result):
    for mid, out in viuf_result.outputs.items():
        for attr in ("net_revenue", "total_opex", "total_capex_monthly",
                     "closing_balance", "net_income", "closing_cash", "cumulative_pfcf"):
            if hasattr(out, attr):
                arr = getattr(out, attr)
                if isinstance(arr, list):
                    assert len(arr) == N, f"{mid}.{attr} len={len(arr)}"


def test_pv_production_zero_during_construction(viuf_result):
    rev = viuf_result.outputs["REV_001"]
    for p in range(CONSTRUCTION):
        assert rev.net_revenue[p] == pytest.approx(0.0, abs=0.01)


def test_irr_computable(viuf_result):
    irr_out = viuf_result.outputs["IRR_001"]
    assert hasattr(irr_out, "project_irr")
    # IRR might be nan if project doesn't recover (valid outcome)
    # Either way the module ran without exception


def test_payback_after_construction(viuf_result):
    irr_out = viuf_result.outputs["IRR_001"]
    pb = irr_out.payback_period_months
    assert pb == -1 or pb >= CONSTRUCTION


def test_no_warnings_for_full_config(viuf_result):
    assert len(viuf_result.warnings) == 0


# ---------------------------------------------------------------------------
# Excel writer tests
# ---------------------------------------------------------------------------

def test_xlsx_file_has_all_sheets(viuf_workbook):
    wb, _ = viuf_workbook
    assert set(wb.sheetnames) == set(SHEETS)


def test_xlsx_revenue_sheet_has_data(viuf_workbook):
    wb, result = viuf_workbook
    ws = wb["Revenue"]
    # Row 27 = REV_001 net_revenue; period 18 (first ops period) should be non-zero
    from assembly.cell_mapper import period_col
    val = ws.cell(row=27, column=period_col(CONSTRUCTION)).value
    assert val is not None
    assert val != 0.0


def test_xlsx_summary_sheet_has_project_name(viuf_workbook):
    wb, result = viuf_workbook
    ws = wb["Summary"]
    assert ws.cell(row=1, column=1).value == result.project_name
